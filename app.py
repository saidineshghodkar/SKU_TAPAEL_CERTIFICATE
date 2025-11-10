"""
Sri Krishnadevaraya University - Tapal Section Data Entry System
"""

from flask import Flask, render_template_string, request, jsonify, send_file
from datetime import datetime, timedelta
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)
DATA_FILE = 'applications_data.json'

CERT_DAYS = {
    'Provisional Certificate': 15,
    'Migration Certificate': 7,
    'Convocation Certificate': 20,
    'Transcript': 15,
    'Genuineness Certificate': 2,
    'Duplicate Marks Memo': 7
}

# UPDATED COLLEGES LIST WITH MULTIPLE TABLE ASSIGNMENTS
COLLEGES_UG = [
    # Table B-1 colleges
    (1, 'Govt. Degree College (Men), Anantapur', 'B-1'),
    (4, 'KH Govt Degree College, Dharmavaram', 'B-1'),
    (6, 'SSS Govt. Degree College, Bukkapatnam', 'B-1'),
    (7, 'PS Govt. Degree College, Penukonda', 'B-1'),
    (9, 'NSPR Govt Degree College for Women, Hindupur', 'B-1'),
    (17, 'Govt. Degree College, Tadipatri', 'B-1'),
    (18, 'SV Arts, Commerce & Science Degree College, Chagala Marri', 'B-1'),
    (20, 'SARM Degree College, Allagadda', 'B-1'),
    (24, 'Govt. Degree College, Nandikotkur', 'B-1'),
    (28, 'Govt. Degree College (Men), Kurnool', 'B-1'),
    (29, 'SJS Degree College (Women), Kurnool', 'B-1'),
    (30, 'SBSYM Degree College, Kurnool', 'B-1'),
    (34, 'Govt Degree College, Yemmiganur', 'B-1'),
    (35, 'The Adoni Arts & Science College, Adoni', 'B-1'),
    (38, 'Government degree Collage ,Dhone', 'B-1'),
    (47, 'CVLNR Cegree Collage,Anantpur', 'B-1'),
    (48, 'MS Degree College, Gooty', 'B-1'),
    (51, 'JCNR Degree Collage , Tadipatri', 'B-1'),
    (56, 'Government degree Collage,Banaganapalli ', 'B-1'),
    (60, 'Govt. Degree College, Srisailam', 'B-1'),
    (63, 'Spandana Degree College, Nandyal', 'B-1'),
    (64, 'National Degree College, Nandyal', 'B-1'),
    (68, 'BVR Degree Collage,Gadivemula ', 'B-1'),
    (71, 'PMR Degree College, Nandyal', 'B-1'),
    (72, 'DRR Degree Collage,Mudigubba ', 'B-1'),
    (73, 'Sri Venkateswara Degree College, Yadiki', 'B-1'),
    (74, 'SNR Degree College, Atmakur', 'B-1'),
    (75, 'Saptagiri Degree College, Hindupur', 'B-1'),
    (76, 'Little Flower Degree College, Anantapur', 'B-1'),
    (77, 'DR Jyothirmayi Degree College, Adoni', 'B-1'),
    (79, 'SPVM Degree College, Adoni', 'B-1'),
    (89, 'Sreenivasa Degree College, Anantapur', 'B-1'),
    (93, 'Sai Degree College, Gannivari Palli, Tadipatri', 'B-1'),
    (94, 'Sri Saraswathy Degree College, Kadiri', 'B-1'),
    (96, 'Sri Vivekananda Degree College, Kurnool', 'B-1'),
    (98, 'Rao s Degree College, Anantapur', 'B-1'),
    (116, 'Basireddy Memorial Degree College, Atmakur', 'B-1'),
    (117, 'Government degree Collage ,Atmakur ', 'B-1'),
    (120, 'Government Degree Collage,Yerraguntla', 'B-1'),
    (127, 'Ravindra Degree College (Vavilala), Kurnool', 'B-1'),
    (131, 'Sri Venkateswara Degree College, Atmakur', 'B-1'),
    (134, 'Sai Sree Degree College, Kanekal, Anantapur Dist', 'B-1'),
    (136, 'Panchamukhi Degree College, Kadiri', 'B-1'),
    (137, 'Sri Vani Degree College, Kadiri', 'B-1'),
    (139, 'Sree Matha Jayalakshmi Degree College, Kadiri', 'B-1'),
    (140, 'Sri Satyakrupa degree Collage,Dharmavaram ', 'B-1'),
    (141, 'Swami Vivekananda Degree Collage,Marepalli, kalyandurg', 'B-1'),
    (143, 'Sri sai Degree collage,Gooty ', 'B-1'),
    (146, 'Sri Sairam degree collage,Codumuru ', 'B-1'),
    (147, 'SRI SATHYASAI DEGREE COLLEGE, KODUMURU', 'B-1'),
    (148, 'SRI VYSHNAVI DEGREE COLLEGE, DHONE', 'B-1'),
    (149, 'SRI KOTI SAI LIPI DEGREE COLLEGE, BATHALAPALLI', 'B-1'),
    (184, 'RMC DEGREE COLLEGE, HALAKUR, AMARAPURAM', 'B-1'),
    (186, 'ANK DEGREE COLLEGE, GORANTLA', 'B-1'),
    (187, 'VIGNAN DEGREE COLLEGE, ODC, KADIRI', 'B-1'),
    (188, 'RUSHI ARTS & SCIENCE DEGREE COLLEGE, KUNDURPI', 'B-1'),
    (190, 'SRI VINAYAKA DEGREE COLLEGE, TALAPULA, KADIRI', 'B-1'),
    (192, 'JYOTHIRMAI DEGREE COLLEGE (WOMEN), KALYANDURG', 'B-1'),
    (195, 'SRI SRINIVASA DEGREE COLLEGE, TANAKALLU', 'B-1'),
    (198, 'SREEDEVI DEGREE COLLEGE, KALYANDURG', 'B-1'),
    (203, 'SRILAKSHMI VIGNESWARA DEGREE COLLEGE, C.K.PALLI', 'B-1'),
    (211, 'SRI VIVEKANANDA DEGREEE COLLEGE, SOMANDEPALLI', 'B-1'),
    (214, 'RAGHAVENDRA INST. OF ARTS & SCI. DC, KR.PALLI CROSS, ANANTAPUR', 'B-1'),
    
    # Table B-2 colleges
    (2, 'SSBN Degree collage ,Anantpur', 'B-2'),
    (3, 'KSN Degree Colage for Women, Anantpur', 'B-2'),
    (5, 'STSN Govt Degree College, Kadiri', 'B-2'),
    (8, 'SDGS Degree College, Hindupur', 'B-2'),
    (9, 'NSPR Govt Degree College for Women, Hindupur', 'B-2'),
    (11, 'SVGM Government Degree Collage,Kalyandurg', 'B-2'),
    (12, 'KTS Govt Degree College, Rayadurg', 'B-2'),
    (14, 'Govt Degree College, Uravakonda', 'B-2'),
    (15, 'SKP Govt Degree College, Guntakal', 'B-2'),
    (16, 'SSGS Degree College, Guntakal', 'B-2'),
    (19, 'SVBD Govt Degree College, Koilkuntla', 'B-2'),
    (21, 'PSC & KVSC Govt Degree College, Nandyal', 'B-2'),
    (23, 'SNSR Arts, Commerce & Science Degree College, Velgode', 'B-2'),
    (25, 'KVR Govt. Degree College for Women, Kurnool', 'B-2'),
    (26, 'Osmania Degree College, Kurnool', 'B-2'),
    (27, 'Osmania Degree College (Women), Kurnool', 'B-2'),
    (31, 'STBC Degree College, Kurnool', 'B-2'),
    (36, 'Govt. Degree College, Alur', 'B-2'),
    (52, 'SPY Reddy Degree College (Women), Nandyal', 'B-2'),
    (57, 'PVKK Degree College, Anantapur', 'B-2'),
    (65, 'Kurnool Degree College, Kurnool', 'B-2'),
    (66, 'Sri Sankara Degree College, Kurnool', 'B-2'),
    (67, 'Smt Teresa Scisarts Degree College, Kurnool', 'B-2'),
    (69, 'KV Subbareddy Degree College, Allagadda', 'B-2'),
    (78, 'Sri Vani Degree College for Women, Anantapur', 'B-2'),
    (82, 'Vasavi Mahila Kalasala, Karnool', 'B-2'),
    (89, 'Sreenivasa Degree College, Anantapur', 'B-2'),
    (118, 'Sri Vivekananda Degree College, kodumuru', 'B-2'),
    (119, 'Sri sivs sai Degree Collage, Mallepalli, Kothacheruvu, Bukkapatnam', 'B-2'),
    (121, 'Srinivasa Degree College, Dharmavaram', 'B-2'),
    (124, 'Sri srisaila Bramarambika DC(Womens),Guntkal', 'B-2'),
    (130, 'Sri Ravi Degree Collage ,Nandyala', 'B-2'),
    (132, 'Sri Venkateswara Degree College, Koilakintla', 'B-2'),
    (137, 'Sri vani Degree College, Kadiri', 'B-2'),
    (138, 'SKNS Amruthavalli Mahila Kalasala, Kadiri', 'B-2'),
    (144, 'S.R. DEGREE COLLEGE, PAMIDI', 'B-2'),
    (195, 'SRI SRINIVASA DEGREE COLLEGE, TANAKALLU', 'B-2'),
    (202, 'SRI VENKATESWARA DEGREE COLLEGE, GUNTAKAL', 'B-2'),
    (204, 'HM DEGREE COLLEGE, GANDLAPENTA', 'B-2'),
    (216, 'SWAMY VIVEKANANDA DEGREE COLLEGE, AKUTHOTHA PALLI, ANANTAPUR', 'B-2'),
    
    # Table B-3 colleges
    (5, 'STSN Govt Degree College, Kadiri', 'B-3'),
    (7, 'PS Govt. Degree College, Penukonda', 'B-3'),
    (26, 'Osmania Degree College, Kurnool', 'B-3'),
    (35, 'The Adoni Arts & Science College, Adoni', 'B-3'),
    (38, 'Government Degree Collage , Dhone', 'B-3'),
    (62, 'St. Joseph Degree College, Kurnool', 'B-3'),
    (65, 'Kurnool Degree College, Kurnool', 'B-3'),
    (66, 'Sri Sankara Degree College, Kurnool', 'B-3'),
    (77, 'Dr Jyothy Ramya Degree zcollage, Adoni', 'B-3'),
    (78, 'Sri Vani Degree College for Women, Anantapur', 'B-3'),
    (79, 'SPVM Degree College, Adoni', 'B-3'),
    (117, 'Govt. Degree College, Atmakur', 'B-3'),
    (120, 'Government Degree Collage, Yerraguntla', 'B-3'),
    (127, 'Ravindra Degree College (Vavilala), Kurnool', 'B-3'),
    (134, 'Sai Sree Degree College, Kanekal, Anantapur Dist', 'B-3'),
    (140, 'Sri Satya Krupa Degree College, Dharmavaram', 'B-3'),
    (143, 'Sri Sai Degree Collage, Gooty', 'B-3'),
    (190, 'Sri Vinayaka Degree Collage,Kadiri', 'B-3'),
    (191, 'Sri Vyshnavi College of Education, Gooty', 'B-3'),
    (194, 'SPACE DEGREE COLLEGE FOR WOMEN, KADIRI', 'B-3'),
    (195, 'Sri Srinivasa Degree Collage,Tanakallu', 'B-3'),
    (197, 'VENKATESWARA DEGREE COLLEGE, NALLACHERUVU', 'B-3'),
    (198, 'SREEDEVI DEGREE COLLEGE, KALYANDURG', 'B-3'),
    (199, 'SRI VIGNAN DEGREE COLLEGE, KADIRI', 'B-3'),
    (201, 'SRI CHAITANYA DEGREE COLLEGE, B.K SAMUNDRAM', 'B-3'),
    (202, 'SRI VENKATESWARA DEGREE COLLEGE, GUNTAKAL', 'B-3'),
    (203, 'SRILAKSHMI VIGNESWARA DEGREE COLLEGE, C.K.PALLI', 'B-3'),
    (204, 'HM DEGREE COLLEGE, GANDLAPENTA', 'B-3'),
    (205, 'ARCHANA DEGREE COLLEGE, NP KUNTA', 'B-3'),
    (206, 'SRI SAI WOMEN\'S DEGREE COLLEGE, GUNTAKAL', 'B-3'),
    (213, 'SRI VASAVI DEGREE COLLEGE, RACHANA PALLI, ANANTAPUR', 'B-3'),
    (226, 'SR Degree Collage, Narpala', 'B-3'),
    (227, 'PENNA COLLEGE OF CEMENT SCIENCE, YADIKI', 'B-3'),
    (304, 'SRI SUBBAIAH DEGREE COLLEGE, ANANTAPUR', 'B-3'),
    
    # Table B-4 colleges
    (1, 'Govt. Degree College (Men), Anantapur', 'B-4'),
    (2, 'SSBN Degree College, Anantapur', 'B-4'),
    (3, 'KSN Degree College for Women, Anantapur', 'B-4'),
    (21, 'PSC & KVSC Govt Degree College, Nandyal', 'B-4'),
    (24, 'Govt. Degree College, Nandikotkur', 'B-4'),
    (64, 'National Degree College, Nandyal', 'B-4'),
    (69, 'KV Subbareddy Degree College, Allagadda', 'B-4'),
    (73, 'Sri Venkateswara Degree College, Yadiki', 'B-4'),
    (116, 'Basi Reddy Memorial Degree College, nandikotkur', 'B-4'),
    (138, 'SKNS Amruthavalli Mahila Kalasala, Kadiri', 'B-4'),
    (144, 'S.R. DEGREE COLLEGE, PAMIDI', 'B-4'),
    (189, 'ENLIGHT COLL.OF SCIENCE & COMMERCE, KODIGINA HALLI, HINDUPUR', 'B-4'),
    (214, 'RAGHAVENDRA INST. OF ARTS & SCI. DC, KR.PALLI CROSS, ANANTAPUR', 'B-4'),
    (215, 'S.L.N. DEGREE COLLEGE, ANANTAPUR', 'B-4'),
    
    # Table B-5 colleges
    (30, 'SBSYM Degree Colage, Kurnool', 'B-5'),
    (32, 'Silver Jubliee Government Degree Collage,Kurnool', 'B-5'),
    (34, 'Govt Degree College, Yemmiganur', 'B-5'),
    (48, 'MS Degree College, Gooty', 'B-5'),
    (51, 'JCNR Degree College, Tadipatri', 'B-5'),
    (56, 'Govt Degree College, Banaganapalli', 'B-5'),
    (63, 'Spandana Degree College, Nandyal', 'B-5'),
    (70, 'TGK institute of Sciences', 'B-5'),
    (77, 'DR Jyothirmayi Degree College, Adoni', 'B-5'),
    (83, 'Sri Venkateswara Degree College, Anantpur', 'B-5'),
    (84, 'Sri Sai Degree College, Kurnool', 'B-5'),
    (85, 'Gowtham Degree College, Anantapur', 'B-5'),
    (93, 'Sai Degree Collage,Nandikotkur', 'B-5'),
    (97, 'Ravi\'s Degree College for Women, Nandyal', 'B-5'),
    (118, 'Sri Vivekananda Degree College, Kodmuru', 'B-5'),
    (137, 'Sri Vani Degree Collage,kadiri', 'B-5'),
    (141, 'SWAMY VIVEKANANDA DEGREE COLLEGE, MAREPALLI, KALYANDURG', 'B-5'),
    (148, 'SRI VYSHNAVI DEGREE COLLEGE, DHONE', 'B-5'),
    (192, 'JYOTHIRMAI DEGREE COLLEGE (WOMEN), KALYANDURG', 'B-5'),
    (193, 'Blue Moon Degree Collage,kadiri', 'B-5'),
    (196, 'MANGALAKARA DEGREE COLLEGE, PUTTAPARTHY', 'B-5'),
    
    # Table B-6 colleges
    (8, 'SDGS Degree College, Hindupur', 'B-6'),
    (9, 'NSPR Govt Degree College for Women, Hindupur', 'B-6'),
    (11, 'SVGM Govt Degree College, Kalyandurg', 'B-6'),
    (12, 'KTS Govt Degree College, Rayadurg', 'B-6'),
    (15, 'SKP Govt Degree College, Guntakal', 'B-6'),
    (16, 'SSGS Degree College, Guntakal', 'B-6'),
    (17, 'Govt. Degree College, Tadipatri', 'B-6'),
     (22, 'sri ramakrishna degree college , nandayala', 'B-6'),
    (29, 'SJS Degree College (Women), Kurnool', 'B-6'),
    (47, 'CVLNR Degree College, Anantapur', 'B-6'),
    (49, 'SYTR Govt. Degree College, Madakasira', 'B-6'),
    (58, 'Master Minds Degree College, Anantapur', 'B-6'),
    (60, 'Govt. Degree College, Srisailam', 'B-6'),
    (63, 'Spandana Degree College, Nandyal', 'B-6'),
    (68, 'BVR Degree college, Gadivemula', 'B-6'),
    (71, 'PMR Degree College, Nandyal', 'B-6'),
    (72, 'DRR Degree College, Mudigubba', 'B-6'),
    (80, 'Balayesu Degree College, Anantapur', 'B-6'),
    (81, 'Sri Balaji Vidya Vihar Degree College, Gorantla', 'B-6'),
    (82, 'Vasavi Mahila Kalasala, Hindupur', 'B-6'),
    (90, 'Sir CV Raman Degree College, Tadipatri', 'B-6'),
    (94, 'Sri Saraswathi Degree college, TADIPATRI', 'B-6'),
    (100, 'Nalanda Degree College, Uravakonda', 'B-6'),
    (122, 'Sri Vivekananda Degree College, Agali', 'B-6'),
     (124, 'Sri Saila bramarimbika Degree college, Guntakal', 'B-6'),
      (125, 'Sri Venkateswara  Degree College,Banganapalli ', 'B-6'),
    (126, 'Sai Degree College, Kurnool', 'B-6'),
    (129, 'Sri Venkateswara Degree College, Nandyal', 'B-6'),
    (139, 'Sree Matha Jayalakshmi Degree College, Kadiri', 'B-6'),
    (142, 'SRI VIDYA DEGREE COLLEGE, KOTHACHERUVU', 'B-6'),
    (149, 'SRI KOTI SAI LIPI DEGREE COLLEGE, BATHALAPALLI', 'B-6'),
    
    
    # Table B-7 colleges
    (4, 'KH Govt Degree College, Dharmavaram', 'B-7'),
    (14, 'Govt Degree College, Uravakonda', 'B-7'),
    (18, 'SV Arts, Commerce & Science Degree College, Chagala Marri', 'B-7'),
    (19, 'SVBD Govt Degree College, Koilkuntla', 'B-7'),
    (20, 'SARM Degree College, Allagadda', 'B-7'),
    (23, 'SNSR Arts, Commerce & Science Degree College, Velgode', 'B-7'),
    (25, 'KVR Govt. Degree College for Women, Kurnool', 'B-7'),
    (27, 'Osmania Degree College (Women), Kurnool', 'B-7'),
    (28, 'Govt. Degree College (Men), Kurnool', 'B-7'),
    (31, 'STBC Degree College, Kurnool', 'B-7'),
    (37, 'Govt. Degree College, Pathhikonda', 'B-7'),
    (52, 'SPY Reddy Degree College (Women), Nandyal', 'B-7'),
    (57, 'PVKK Degree College, Anantapur', 'B-7'),
    (59, 'Sri Vijayadurg Degree College, Kurnool', 'B-7'),
    (74, 'SNR Degree College, Atmakur', 'B-7'),
    (75, 'Saptagiri Degree College, Hindupur', 'B-7'),
    (76, 'Little Flower Degree College, Anantapur', 'B-7'),
    (89, 'Sreenivasa Degree College, Anantapur', 'B-7'),
    (92, 'Royal PG College, Nandikotkur', 'B-7'),
    (97, 'Ravi\'s Degree College for Women, Nandyal', 'B-7'),
    (98, 'Rao\'s Degree College, Anantapur', 'B-7'),
    (99, 'Sri Sai\'s Womens Degree College, Anantapur', 'B-7'),
    (115, 'Mahatma Degree College, Nandikotkur', 'B-7'),
    (119, 'Sri Siva Sai Degree College, Yerraguntla', 'B-7'),
    (121, 'Srinivasa Ddgree college, Dharmavaram', 'B-7'),
    (128, 'Vibhav Degree College, Yellanur Road, Tadpatri', 'B-7'),
    (130, 'Sri Ravi Degree College, Atmakur', 'B-7'),
    (131, 'Sri Venkateswara Degree College, Koilkuntla', 'B-7'),
    (132, 'Sri Venkateswara Degree College, Yemmiganur', 'B-7'),
    (133, 'Vyshnavi Degree College, Dhone', 'B-7'),
    (145, 'SRI SATYA SAI DEGREE COLLEGE, PENUKONDA', 'B-7'),
    (146, 'SAI RAM DEGREE COLLEGE, KODUMURU', 'B-7'),
    (150, 'SRI SAI KRUPA DEGREE COLLEGE, DHARMAVARAM', 'B-7'),
    (185, 'SRI SAI DEGREE COLLEGE, DHARMAVARAM', 'B-7'),
    (186, 'ANK DEGREE COLLEGE, GORANTLA', 'B-7'),
    (187, 'VIGNAN DEGREE COLLEGE, ODC, KADIRI', 'B-7'),
    (211, 'SRI VIVEKANANDA DEGREEE COLLEGE, SOMANDEPALLI', 'B-7'),
    (212, 'SRI VIVEKANANDA DEGREE COLLEGE, ANANTAPUR', 'B-7'),
    (216, 'SWAMY VIVEKANANDA DEGREE COLLEGE, AKUTHOTHA PALLI, ANANTAPUR', 'B-7'),
    (217, 'SREE VIGNAN DEGREE COLLEGE, AMADAGURU', 'B-7'),
    (225, 'SRI VENKATESWARA DEGREE COLLEGE, MADAKASIRA', 'B-7'),
    
    # Table B-8 colleges
    (4, 'KH Govt Degree College, Dharmavaram', 'B-8'),
    (18, 'SV Arts, Commerce & Science Degree College, Chagala Marri', 'B-8'),
    (19, 'SVBD Govt Degree College, Koilkuntla', 'B-8'),
    (20, 'SARM Degree College, Allagadda', 'B-8'),
    (24, 'Govt. Degree College, Nandikotkur', 'B-8'),
    (25, 'KVR Govt. Degree College for Women, Kurnool', 'B-8'),
    (26, 'Osmania Degree College, Kurnool', 'B-8'),
    (27, 'Osmania Degree College (Women), Kurnool', 'B-8'),
    (28, 'Govt. Degree College (Men), Kurnool', 'B-8'),
    (31, 'STBC Degree College, Kurnool', 'B-8'),
    (32, 'Silver Jubilee Govt Degree College, Kurnool', 'B-8'),
    (35, 'The Adoni Arts & Science College, Adoni', 'B-8'),
    (62, 'St. Joseph Degree College, Kurnool', 'B-8'),
    (65, 'Kurnool Degree College, Kurnool', 'B-8'),
    (67, 'Smt Teresa Scisarts Degree College, Kurnool', 'B-8'),
    (74, 'SNR Degree College, Atmakur', 'B-8'),
    (81, 'Sri Balaji Vidya Vihar Degree College, Gorantla', 'B-8'),
    (83, 'Sri Venkateswara Degree College, Hindupur', 'B-8'),
    (84, 'Sri Sai Degree College, Kurnool', 'B-8'),
    (85, 'Gowtham Degree College, Anantapur', 'B-8'),
    (93, 'Sai Degree College, Gannivari Palli, Tadipatri', 'B-8'),
    (99, 'Sri Sai\'s Womens Degree College, Anantapur', 'B-8'),
    (115, 'Mahatma Degree College, Nandikotkur', 'B-8'),
    (116, 'Basireddy Memorial Degree College, Atmakur', 'B-8'),
    (118, 'Sri Vivekananda Degree College, Mallepalli, Kothacheruvu, Bukkapatnam', 'B-8'),
    (119, 'Sri Siva Sai Degree College, Yerraguntla', 'B-8'),
    (121, 'Srinivasa Degree College, Dharmavaram', 'B-8'),
    (122, 'Sri Vivekananda Degree College, Agali', 'B-8'),
    (130, 'Sri Ravi Degree College, Atmakur', 'B-8'),
    (136, 'Panchamukhi Degree College, Kadiri', 'B-8'),
    (139, 'Sree Matha Jayalakshmi Degree College, Kadiri', 'B-8'),
    (141, 'SWAMY VIVEKANANDA DEGREE COLLEGE, MAREPALLI, KALYANDURG', 'B-8'),
    (142, 'SRI VIDYA DEGREE COLLEGE, KOTHACHERUVU', 'B-8'),
    
    
    # Table B-9 colleges
    (1, 'Govt. Degree College (Men), Anantapur', 'B-9'),
    (2, 'SSBN Degree College, Anantapur', 'B-9'),
    (3, 'KSN Degree College for Women, Anantapur', 'B-9'),
    (9, 'NSPR Govt Degree College for Women, Hindupur', 'B-9'),
    (11, 'SVGM Govt Degree College, Kalyandurg', 'B-9'),
    (12, 'KTS Govt Degree College, Rayadurg', 'B-9'),
    (47, 'CVLNR Degree College, Anantapur', 'B-9'),
    (51, 'JCNR Degree College, Tadipatri', 'B-9'),
    (69, 'KV Subbareddy Degree College, Allagadda', 'B-9'),
    (82, 'Vasavi Mahila Kalasala, Hindupur', 'B-9'),
    (89, 'Sreenivasa Degree College, Anantapur', 'B-9'),
    (90, 'Sir CV Raman Degree College, Tadipatri', 'B-9'),
    (92, 'Royal PG College, Nandikotkur', 'B-9'),
    (94, 'Sri Saraswathy Degree College, Kadiri', 'B-9'),
    (117, 'Govt. Degree College, Kodumuru', 'B-9'),
    (120, 'Govt. Degree College, Dharmavaram', 'B-9'),
    (125, 'Sri Venkateswara Degree College, Adoni', 'B-9'),
    (127, 'Ravindra Degree College (Vavilala), Kurnool', 'B-9'),
    (128, 'Vibhav Degree College, Yellanur Road, Tadpatri', 'B-9'),
    (132, 'Sri Venkateswara Degree College, Yemmiganur', 'B-9'),
    (134, 'Sai Sree Degree College, Kanekal, Anantapur Dist', 'B-9'),
    (138, 'SKNS Amruthavalli Mahila Kalasala, Kadiri', 'B-9'),
    (147, 'SRI SATHYASAI DEGREE COLLEGE, KODUMURU', 'B-9'),
    (190, 'SRI VINAYVAKA DEGREE COLLEGE, KADIRI', 'B-9'),
    (211, 'SRI VIVAKANADHA DEGREE COLLEGE, SOMANDEPALLI', 'B-9'),
    (212, 'SRI VINAYVAKA DEGREE COLLEGE, ANANTAPUR', 'B-9'),
    (213, 'SRI VASAVI DEGREE COLLEGE, ANANTAPUR', 'B-9'),
    (214, 'RAGHAVENDRA INST. OF ARTS & SCI. DC, KR.PALLI CROSS, ANANTAPUR', 'B-9'),
    (215, 'S.L.N. DEGREE COLLEGE, ANANTAPUR', 'B-9'),
    (216, 'SWAMY VIVEKANANDA DEGREE COLLEGE, AKUTHOTHA PALLI, ANANTAPUR', 'B-9'),
    (225, 'SRI VENKATESWARA DEGREE COLLEGE, MADAKASIRA', 'B-9'),
    (226, 'S.R. DEGREE COLLEGE, NARPALA', 'B-9'),
    (304, 'SRI SUBBAIAH DEGREE COLLEGE, ANANTAPUR', 'B-9'),
    
    # Table B-10 colleges
    (7, 'PS Govt. Degree College, Penukonda', 'B-10'),
    (8, 'SDGS Degree College, Hindupur', 'B-10'),
    (15, 'SKP Govt Degree College, Guntakal', 'B-10'),
    (16, 'SSGS Degree College, Guntakal', 'B-10'),
    (49, 'SYTR Govt. Degree College, Madakasira', 'B-10'),
    (72, 'DRR Degree College, Mudigubba', 'B-10'),
    (79, 'SPVM Degree College, Adoni', 'B-10'),
    (80, 'Balayesu Degree College, Anantapur', 'B-10'),
    (123, 'Vidya Ratna Dr KGA Gupta Degree College, Guntakal', 'B-10'),
    (124, 'Srisaila Bramarambika Degree College (Womens), Banaganapalli', 'B-10'),
    (135, 'Sonia Gandhi Degree College, Kanekal, Anantapur Dist', 'B-10'),
    (144, 'S.R. DEGREE COLLEGE, PAMIDI', 'B-10'),
    (184, 'RMC DEGREE COLLEGE,AMARAPURAM', 'B-10'),
    (186, 'ANK DEGREE COLLEGE, GORANTLA', 'B-10'),
    (187, 'VIGNAN DEGREE COLLEGE, ODC, KADIRI', 'B-10'),
    (188, 'RUSHI ARTS & SCIENCE DEGREE COLLEGE, KUNDURPI', 'B-10'),
    (189, 'ENLIGHT COLL.OF SCIENCE & COMMERCE, KODIGINA HALLI, HINDUPUR', 'B-10'),
    (193, 'BLUEMOON DEGREE COLLEGE, KADIRI', 'B-10'),
    (194, 'SPACE DEGREE COLLEGE FOR WOMEN, KADIRI', 'B-10'),
    (195, 'SRI SRINIVASA DEGREE COLLEGE, TANAKALLU', 'B-10'),
    (196, 'MANGALAKARA DEGREE COLLEGE, PUTTAPARTHY', 'B-10'),
    (197, 'VENKATESWARA DEGREE COLLEGE, NALLACHERUVU', 'B-10'),
    (198, 'SREEDEVI DEGREE COLLEGE, KALYANDURG', 'B-10'),
    (199, 'SRI VIGNAN DEGREE COLLEGE, KADIRI', 'B-10'),
    (201, 'SRI CHAITANYA DEGREE COLLEGE, B.K SAMUNDRAM', 'B-10'),
     (202, 'SRI VENKATESWARA DEGREE COLLEGE, GUNTAKAL', 'B-10'),
    (203, 'SRILAKSHMI VIGNESWARA DEGREE COLLEGE, C.K.PALLI', 'B-10'),
    (204, 'HM  D.C., GANDLAPENTA', 'B-10'),
    (205, 'ARCHANA D.C., NP KUNTA', 'B-10'),
    (206, 'SRI SAI WONMENS D.C.,GUNTKAL', 'B-10'),
    (207, 'SRI VIVEKANTA COLLEGE OF ARTS & SCIENCE, CHILANATHUR', 'B-10'),
    (208, 'SRI KOTIRAMADU DEGREE COLLEGE,TADIMARRI', 'B-10'),
    (209, 'SRI SARADA DEGREE COLLEGE, GANDLAPENTA', 'B-10'),
    (210, 'KURLI ANNAPURNAMMA RAMIREDDY D.C., NALLAMADA', 'B-10'),
    
    # Table B-11 colleges
    (5, 'STSN Govt Degree College, Kadiri', 'B-11'),
    (6, 'SSS Govt. Degree College, Bukkapatnam', 'B-11'),
    (14, 'Govt Degree College, Uravakonda', 'B-11'),
    (17, 'Govt. Degree College, Tadipatri', 'B-11'),
    (21, 'PSC & KVSC Govt Degree College, Nandyal', 'B-11'),
    (22, 'PSC AND KVSC GOVT DEGREE COLLEGE NANDYAL ', 'B-11'),
    (23, 'SNSR Arts, Commerce & Science Degree College, Velgode', 'B-11'),
    (29, 'SJS Degree College (Women), Kurnool', 'B-11'),
    (30, 'SBSYM Degree College, Kurnool', 'B-11'),
    (34, 'Govt Degree College, Yemmiganur', 'B-11'),
    (37, 'Govt. Degree College, Pathhikonda', 'B-11'),
    (38, 'Govt. Degree College, Dhone', 'B-11'),
    (48, 'MS Degree College, Gooty', 'B-11'),
    (52, 'SPY Reddy Degree College (Women), Nandyal', 'B-11'),
    (56, 'Govt Degree College, Banaganapalli', 'B-11'),
    (57, 'PVKK Degree College, Anantapur', 'B-11'),
    (58, 'Master Minds Degree College, Anantapur', 'B-11'),
    (59, 'Sri Vijayadurg Degree College, Kurnool', 'B-11'),
    (60, 'Govt. Degree College, Srisailam', 'B-11'),
    (63, 'Spandana Degree College, Nandyal', 'B-11'),
    (64, 'National Degree College, Nandyal', 'B-11'),
    (66, 'Sri Sankara Degree College, Kurnool', 'B-11'),
    (68, 'BVR Degree College, Gadivemula', 'B-11'),
    (71, 'PMR Degree College, Nandyal', 'B-11'),
    (73, 'Sri Venkateswara Degree College, Yadiki', 'B-11'),
    (75, 'Saptagiri Degree College, Hindupur', 'B-11'),
    (76, 'Little Flower Degree College, Anantapur', 'B-11'),
    (77, 'DR Jyothirmayi Degree College, Adoni', 'B-11'),
    (78, 'Sri Vani Degree College for Women, Anantapur', 'B-11'),
    (96, 'Sri Vivekananda Degree College, Kurnool', 'B-11'),
    (98, 'Rao Degree College, Anantapur', 'B-11'),
    (100, 'Nalanda Degree College, Uravakonda', 'B-11'),
    (126, 'Sai Degree College, Adoni', 'B-11'),
    (129, 'Sri Venkateswara Degree College, Nandyal', 'B-11'),
    (131, 'Sri Venkateswara Degree College, Koilkuntla', 'B-11'),
    (133, 'Vyshnavi Degree College, Dhone', 'B-11'),
    (140, 'Sri Satya Krupa Degree College, Dharmavaram', 'B-11'),
    (143, 'SRI SAI DEGREE COLLEGE, GOOTY', 'B-11'),
    (145, 'SRI SATYA SAI DEGREE COLLEGE, PENUKONDA', 'B-11'),
    (148, 'SRI VYSHNAVI DEGREE COLLEGE, DHONE', 'B-11'),
    (149, 'SRI KOTI SAI LIPI DEGREE COLLEGE, BATHALAPALLI', 'B-11'),
    (150, 'SRI SAI KRUPA DEGREE COLLEGE, DHARMAVARAM', 'B-11'),
    (185, 'SRI SAI DEGREE COLLEGE, DHARMAVARAM', 'B-11'),
    (192, 'JYOTHIRMAI DEGREE COLLEGE (WOMEN), KALYANDURG', 'B-11'),
    
    # Table B-12 colleges (B.Ed. colleges)
  
    (86, 'Sri Sai Ram College of Education, Ramapuram, Penukonda', 'B-12'),
    (87, 'Haindavi College of Education, Dharmavaram', 'B-12'),
    (91, 'Sri Balaji College of Education, Kothacheruvu', 'B-12'),
    (95, 'crecent  of Education, Anathapuram ,', 'B-12'),
    (101, 'Intel Institute of Teacher Training College of Education, Anantapur', 'B-12'),
    (102, 'Sri Balaji College of Education, Anantapur', 'B-12'),
    (103, 'little flower college of education for women,ananpur', 'B-12'),
    (105, 'KC Narayana College of Education, Guntakal', 'B-12'),
    (110, 'St. Joseph College of Education, Rachanapalli, Anantapur', 'B-12'),
    (111, 'St. Joseph College of Education, MPR Dam, Anantapur', 'B-12'),
    (112, 'Hailee College of Education, Kalyandurg', 'B-12'),
    (151, 'S.K.U. College of Education, S.K. University, Anantapur', 'B-12'),
    (152, 'S.V. College of Education, Anantapur', 'B-12'),
    (153, 'Suseela College of Education, Anantapur', 'B-12'),
    (154, 'A.M. Linganna College of Education, Hindupur', 'B-12'),
    (155, 'St. Xavier\'s College of Education, Hindupur', 'B-12'),
    (156, 'Dr. B.V. Syamala Ratnam College of Education, Hindupur', 'B-12'),
    (157, 'Venkateswara College of Education, Kadiri', 'B-12'),
    (158, 'Sri Devi College of Education, Borampalli, Kalyandurg', 'B-12'),
    (159, 'Swamy Vivekananda College of Education, Kalyandurg', 'B-12'),
    (160, 'Sir C.V. Raman College of Education, Tadipatri', 'B-12'),
    (161, 'Sai Siddhartha College of Education, Tadipatri', 'B-12'),
    (182, 'Sri Saibaba College of Education, Anantapuramu', 'B-12'),
    (191, 'Sri Vyshnavi College of Education, Gooty', 'B-12'),
    (217, 'Vagdevi College of Education, Yadiki', 'B-12'),
    (218, 'Deekshitha B.Ed. College, Tadipatri', 'B-12'),
    (220, 'Sri Vidyanikhetan College of Education, Tadipatri', 'B-12'),
    (221, 'Narayana College of Education, Tadipatri', 'B-12'),
    (222, 'Rana B.Ed. College of Education, Penukonda', 'B-12'),
    (223, 'Sree Vidyaranya B.P.Ed. College, Kurnool Rd., Gooty', 'B-12'),
    (224, 'Narayana B.P.Ed. College of Physical Education, Tadipatri', 'B-12'),
    (228, 'Sri Vijayabharathi College of Education, Bukkarayasamudram', 'B-12'),
]

# UPDATED REQUIRED DOCS WITH NEW FIELDS
REQUIRED_DOCS = {
    'Duplicate Marks Memo': ['SBI Bank Challan', 'Application Form', 'Marks Memos', 'Cloth Cover', 'Postal Stamp', 'Aadhar Card', 'APPAR ID'],
    'Provisional Certificate': ['SBI Bank Challan', 'Application Form', 'Marks Memos', 'Convocation Certificate', 'Migration Certificate', 'Intermediate Marks', '10th Marks', 'Aadhar Card', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
    'Convocation Certificate': ['SBI Bank Challan', 'Application Form', 'CCM & Provisional', 'Convocation Certificate', 'Intermediate Marks', '10th Marks', 'Aadhar Card', 'Cloth Cover', 'Photos', 'Gazetted Signature', 'APPAR ID'],
    'Transcript': ['SBI Bank Challan', 'Application Form', 'CCM & Provisional', 'Certificates', 'Cloth Cover', 'Postal Stamp', 'Aadhar Card', 'APPAR ID'],
    'Migration Certificate': ['SBI Bank Challan', 'Application Form', 'Degree Certificate', 'Marks Memos', 'Provisional Certificate', 'Aadhar Card', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
    'Genuineness Certificate': ['Notification', 'Proceeding Copy', 'Journals', 'Marks Memos', 'Convocation Certificate', 'Aadhar Card', 'Photos', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
    'Duplicate Certificate': ['SBI Bank Challan', 'Application Form', 'Original Certificate', 'Aadhar Card', 'Photos', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
    'Mphill/Phd Provisional certificate': ['Notification', 'Proceeding Copy', 'Journals', 'Marks Memos', 'Convocation Certificate', 'Aadhar Card', 'Photos', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
    'Mphill/Phd Convocation certificate': ['Notification', 'Proceeding Copy', 'Journals', 'Marks Memos', 'MPhil Provisional', 'Convocation Certificate', 'Aadhar Card', 'Photos', 'Gazetted Signature', 'Cloth Cover', 'Postal Stamp', 'APPAR ID']
}

def init_db():
    if not os.path.exists(DATA_FILE):
        data = {'applications': [], 'dispatches': [], 'metadata': {'created': datetime.now().isoformat()}}
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

def load_data():
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        init_db()
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def calc_status(submitted_at, cert_type):
    try:
        submitted = datetime.fromisoformat(submitted_at.replace('Z', '+00:00').split('.')[0])
        days_needed = CERT_DAYS.get(cert_type, 15)
        days_passed = (datetime.now() - submitted).days
        if days_passed < (days_needed * 0.5):
            return 'within_time'
        elif days_passed < days_needed:
            return 'out_of_time'
        else:
            return 'overdue'
    except:
        return 'pending'

@app.route('/')
def index():
    init_db()
    return render_template_string(HTML)

@app.route('/api/colleges')
def get_colleges():
    return jsonify([{'code': str(c[0]), 'name': c[1], 'table': c[2]} for c in COLLEGES_UG])

@app.route('/check-hallticket', methods=['POST'])
def check_hallticket():
    data = request.json
    hallticket = data.get('hallticket', '').strip()
    db = load_data()
    for app in db.get('applications', []):
        if app.get('formData', {}).get('hallticket') == hallticket:
            return jsonify({'exists': True})
    return jsonify({'exists': False})

@app.route('/check-dispatch', methods=['POST'])
def check_dispatch():
    data = request.json
    hallticket = data.get('hallticket', '').strip()
    db = load_data()
    
    # Get all pending applications for this hallticket
    pending_apps = []
    for app in db.get('applications', []):
        if app.get('formData', {}).get('hallticket') == hallticket and app.get('status') != 'closed':
            pending_apps.append(app)
    
    if pending_apps:
        return jsonify({'exists': True, 'pending_apps': len(pending_apps)})
    return jsonify({'exists': False, 'pending_apps': 0})

@app.route('/save-application', methods=['POST'])
def save_app():
    try:
        data = request.json
        form_type = data.get('formType')
        form_data = data.get('data', {})
        
        # Validate required fields
        if not form_type:
            return jsonify({'success': False, 'error': 'Form type is required'}), 400
            
        if not form_data.get('hallticket'):
            return jsonify({'success': False, 'error': 'Hall ticket is required'}), 400
        
        db = load_data()
        next_id = max([a.get('id', 0) for a in db.get('applications', [])], default=0) + 1
        
        new_app = {
            'id': next_id,
            'formType': form_type,
            'timestamp': datetime.now().isoformat(),
            'submittedAt': datetime.now().isoformat(),
            'formData': form_data,
            'status': 'pending'
        }
        
        db['applications'].append(new_app)
        save_data(db)
        return jsonify({'success': True, 'id': next_id})
        
    except Exception as e:
        print(f"ERROR in save_app: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/save-dispatch', methods=['POST'])
def save_dispatch():
    try:
        data = request.json
        hallticket = data.get('hallticket', '').strip()
        mode = data.get('dispatchMode', '')
        app_id = data.get('app_id')  # Get specific application ID
        
        print(f"DEBUG: Saving dispatch for hallticket: {hallticket}, mode: {mode}, app_id: {app_id}")
        
        if not hallticket or not mode:
            return jsonify({'success': False, 'error': 'Missing hall ticket or dispatch mode'}), 400
        
        db = load_data()
        
        # Find the specific application
        app_found = None
        app_data = None
        for app in db.get('applications', []):
            app_form_data = app.get('formData', {})
            if app_form_data and app_form_data.get('hallticket') == hallticket:
                if app_id:  # If specific app_id provided, match it
                    if app.get('id') == app_id:
                        app_found = app
                        app_data = app_form_data
                        break
                else:  # Otherwise take first pending application
                    if app.get('status') != 'closed':
                        app_found = app
                        app_data = app_form_data
                        break
        
        if not app_found:
            return jsonify({'success': False, 'error': 'No pending application found for this hall ticket'}), 404
        
        # Check if this specific application already has a dispatch
        for dispatch in db.get('dispatches', []):
            if dispatch.get('app_id') == app_found.get('id') and dispatch.get('status') != 'completed':
                return jsonify({'success': False, 'error': 'This application is already being processed'}), 400
        
        # Create dispatch record
        dispatch_record = {
            'id': len(db.get('dispatches', [])) + 1,
            'app_id': app_found.get('id'),  # Store application ID
            'hallticket': hallticket,
            'dispatchMode': mode,
            'timestamp': datetime.now().isoformat(),
            'status': 'processed',
            'applicationData': app_data
        }
        
        db['dispatches'].append(dispatch_record)
        save_data(db)
        
        print(f"DEBUG: Dispatch processed successfully for {hallticket}, app_id: {app_found.get('id')}")
        return jsonify({'success': True, 'message': 'Dispatch processed successfully'})
        
    except Exception as e:
        print(f"ERROR in save_dispatch: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/close-dispatch', methods=['POST'])
def close_dispatch():
    try:
        data = request.json
        hallticket = data.get('hallticket', '').strip()
        tracking_id = data.get('tracking_id', '').strip()
        app_id = data.get('app_id')  # Get specific application ID
        
        print(f"DEBUG: Closing dispatch for hallticket: {hallticket}, tracking: {tracking_id}, app_id: {app_id}")
        
        if not hallticket or not tracking_id:
            return jsonify({'success': False, 'error': 'Missing hall ticket or tracking ID/reason'}), 400
        
        db = load_data()
        
        # Find the processed dispatch for this specific application
        dispatch_found = None
        for dispatch in db.get('dispatches', []):
            if dispatch.get('hallticket') == hallticket:
                if app_id:  # If specific app_id provided, match it
                    if dispatch.get('app_id') == app_id:
                        dispatch_found = dispatch
                        break
                else:  # Otherwise take first processed dispatch
                    if dispatch.get('status') == 'processed':
                        dispatch_found = dispatch
                        break
        
        if not dispatch_found:
            return jsonify({'success': False, 'error': 'No processed dispatch found for this hall ticket'}), 404
        
        if dispatch_found.get('status') == 'completed':
            return jsonify({'success': False, 'error': 'Dispatch already closed'}), 400
        
        # Update dispatch to completed status
        dispatch_found['status'] = 'completed'
        dispatch_found['tracking_id'] = tracking_id
        dispatch_found['closed_at'] = datetime.now().isoformat()
        
        # Update specific application status to closed
        app_found = False
        for app in db.get('applications', []):
            if app.get('id') == dispatch_found.get('app_id'):
                app['status'] = 'closed'
                app_found = True
                break
        
        if not app_found:
            return jsonify({'success': False, 'error': 'Application not found'}), 404
        
        save_data(db)
        
        print(f"DEBUG: Dispatch closed successfully for {hallticket}, app_id: {dispatch_found.get('app_id')}")
        return jsonify({'success': True, 'message': 'Dispatch closed successfully and application marked as closed'})
        
    except Exception as e:
        print(f"ERROR in close_dispatch: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-pending-applications', methods=['POST'])
def get_pending_applications():
    try:
        data = request.json
        hallticket = data.get('hallticket', '').strip()
        
        if not hallticket:
            return jsonify({'success': False, 'error': 'Hall ticket is required'}), 400
        
        db = load_data()
        pending_apps = []
        
        for app in db.get('applications', []):
            app_form_data = app.get('formData', {})
            if app_form_data and app_form_data.get('hallticket') == hallticket and app.get('status') != 'closed':
                # Check if this app already has a completed dispatch
                has_completed_dispatch = False
                for dispatch in db.get('dispatches', []):
                    if dispatch.get('app_id') == app.get('id') and dispatch.get('status') == 'completed':
                        has_completed_dispatch = True
                        break
                
                if not has_completed_dispatch:
                    pending_apps.append({
                        'id': app.get('id'),
                        'formType': app.get('formType'),
                        'name': app_form_data.get('name', 'N/A'),
                        'certificateType': app_form_data.get('certificateType', 'N/A'),
                        'submittedAt': app.get('submittedAt'),
                        'status': app.get('status')
                    })
        
        return jsonify({'success': True, 'pending_apps': pending_apps})
        
    except Exception as e:
        print(f"ERROR in get_pending_applications: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-processed-dispatches', methods=['POST'])
def get_processed_dispatches():
    try:
        data = request.json
        hallticket = data.get('hallticket', '').strip()
        
        if not hallticket:
            return jsonify({'success': False, 'error': 'Hall ticket is required'}), 400
        
        db = load_data()
        processed_dispatches = []
        
        for dispatch in db.get('dispatches', []):
            if dispatch.get('hallticket') == hallticket and dispatch.get('status') == 'processed':
                # Find the application details
                app_data = None
                for app in db.get('applications', []):
                    if app.get('id') == dispatch.get('app_id'):
                        app_data = app.get('formData', {})
                        break
                
                processed_dispatches.append({
                    'id': dispatch.get('id'),
                    'app_id': dispatch.get('app_id'),
                    'dispatchMode': dispatch.get('dispatchMode'),
                    'timestamp': dispatch.get('timestamp'),
                    'name': app_data.get('name', 'N/A') if app_data else 'N/A',
                    'certificateType': app_data.get('certificateType', 'N/A') if app_data else 'N/A'
                })
        
        return jsonify({'success': True, 'processed_dispatches': processed_dispatches})
        
    except Exception as e:
        print(f"ERROR in get_processed_dispatches: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-all-data')
def get_all():
    db = load_data()
    for app in db.get('applications', []):
        app['calculatedStatus'] = calc_status(app.get('submittedAt'), app.get('formData', {}).get('certificateType', ''))
    return jsonify(db)

@app.route('/get-dispatches')
def get_dispatches():
    try:
        db = load_data()
        return jsonify({'success': True, 'dispatches': db.get('dispatches', [])})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete-application', methods=['POST'])
def delete_app():
    try:
        app_id = request.json.get('id')
        db = load_data()
        db['applications'] = [a for a in db.get('applications', []) if a.get('id') != app_id]
        save_data(db)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export-excel', methods=['POST'])
def export_excel():
    try:
        data = request.json
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        table_num = data.get('tableNumber')
        
        db = load_data()
        apps = db.get('applications', [])
        
        if from_date and to_date:
            from_dt = datetime.fromisoformat(from_date)
            to_dt = datetime.fromisoformat(to_date)
            apps = [a for a in apps if from_dt <= datetime.fromisoformat(a['submittedAt'].split('.')[0]) <= to_dt]
        
        if table_num:
            apps = [a for a in apps if a.get('formData', {}).get('tableNumber') == table_num]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        
        headers = ['S.No', 'Form Type', 'Name', 'Register', 'Hall Ticket', 'Application Details/Address', 'Degree', 'Course', 'Table', 'Certificate', 'Date', 'Status']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for idx, app in enumerate(apps, 1):
            fd = app.get('formData', {})
            submitted = datetime.fromisoformat(app.get('submittedAt', '').split('.')[0])
            status = app.get('status', 'pending').upper()
            
            if app.get('status') == 'pending':
                status = calc_status(app.get('submittedAt'), fd.get('certificateType', '')).replace('_', ' ').upper()
            
            ws.append([
                idx, app.get('formType'), fd.get('name'), fd.get('register'),
                fd.get('hallticket'), 
                fd.get('address') or fd.get('applicationDetails') or '',  # Address for regular/distance, applicationDetails for confidential
                fd.get('degree'), fd.get('course'), fd.get('tableNumber'),
                fd.get('certificateType'), submitted.strftime('%Y-%m-%d'), status
            ])
        
        for column in ws.columns:
            max_len = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    except Exception as e:
        print(f"ERROR in export_excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sri Krishnadevaraya University - Tapal Section</title>
    <style>
        * { 
            margin: 0; 
            padding: 0; 
            box-sizing: border-box; 
        }
        
        body { 
            font-family: 'Segoe UI', 'Inter', 'Roboto', sans-serif; 
            background: linear-gradient(135deg, #f5f7fa 0%, #e4efe9 100%);
            color: #2c3e50; 
            line-height: 1.6;
        }
        
        header { 
            background: linear-gradient(135deg, #1a2f6b 0%, #2a5298 100%); 
            padding: 25px 30px; 
            box-shadow: 0 8px 25px rgba(0,0,0,0.15); 
            display: flex; 
            align-items: center; 
            gap: 25px; 
            border-bottom: 5px solid #ffd700;
            position: relative;
            overflow: hidden;
        }
        
        header::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: url('data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100" opacity="0.05"><text x="50%" y="50%" font-size="20" text-anchor="middle" fill="white">SKU</text></svg>');
            pointer-events: none;
        }
        
        header img { 
            width: 80px; 
            height: 80px; 
            border-radius: 12px; 
            border: 3px solid #ffd700; 
            object-fit: cover; 
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
            background: white;
            padding: 5px;
        }
        
        .header-content {
            flex: 1;
        }
        
        header h1 { 
            color: white; 
            font-size: 28px; 
            font-weight: 700; 
            line-height: 1.2;
            text-shadow: 0 2px 4px rgba(0,0,0,0.3);
            margin-bottom: 5px;
        }
        
        header p { 
            color: #e8f4f8; 
            font-size: 16px; 
            opacity: 0.95;
            font-weight: 500;
        }
        
        .nav-tabs { 
            display: flex; 
            background: white; 
            border-bottom: 3px solid #2a5298; 
            padding: 0 30px; 
            gap: 40px; 
            box-shadow: 0 4px 15px rgba(0,0,0,0.08);
            position: relative;
        }
        
        .nav-tab { 
            padding: 18px 25px; 
            background: none; 
            border: none; 
            font-size: 16px; 
            font-weight: 600; 
            color: #7f8c8d; 
            cursor: pointer; 
            border-bottom: 4px solid transparent; 
            transition: all 0.4s cubic-bezier(0.25, 0.46, 0.45, 0.94);
            position: relative;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        
        .nav-tab::before {
            content: '';
            position: absolute;
            bottom: -3px;
            left: 50%;
            width: 0;
            height: 3px;
            background: #2a5298;
            transition: all 0.4s cubic-bezier(0.25, 0.46, 0.45, 0.94);
            transform: translateX(-50%);
        }
        
        .nav-tab.active { 
            color: #2a5298; 
        }
        
        .nav-tab.active::before {
            width: 100%;
        }
        
        .nav-tab:hover { 
            color: #2a5298; 
            transform: translateY(-2px);
        }
        
        .container { 
            max-width: 1400px; 
            margin: 40px auto; 
            padding: 0 25px; 
        }
        
        .section { 
            display: none; 
            animation: fadeIn 0.6s ease-in-out;
        }
        
        .section.active { 
            display: block; 
        }
        
        .card { 
            background: white; 
            border-radius: 16px; 
            box-shadow: 0 8px 30px rgba(0,0,0,0.12); 
            padding: 40px; 
            margin-bottom: 30px; 
            border: 1px solid #eef2f7;
            transition: transform 0.3s ease, box-shadow 0.3s ease;
            position: relative;
            overflow: hidden;
        }
        
        .card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            width: 5px;
            height: 100%;
            background: linear-gradient(135deg, #2a5298 0%, #1a2f6b 100%);
        }
        
        .card:hover {
            transform: translateY(-5px);
            box-shadow: 0 15px 40px rgba(0,0,0,0.15);
        }
        
        .form-group { 
            margin-bottom: 25px; 
            position: relative;
        }
        
        .form-group label { 
            display: block; 
            margin-bottom: 10px; 
            font-weight: 600; 
            color: #2c3e50;
            font-size: 15px;
        }
        
        .form-group.required label::after { 
            content: ' *'; 
            color: #e74c3c; 
        }
        
        input, select, textarea { 
            width: 100%; 
            padding: 14px 16px; 
            border: 2px solid #e0e6ed; 
            border-radius: 10px; 
            font-size: 15px; 
            transition: all 0.3s ease;
            background: #fafbfc;
            font-family: inherit;
        }
        
        input:focus, select:focus, textarea:focus { 
            border-color: #2a5298; 
            outline: none; 
            box-shadow: 0 0 0 4px rgba(42, 82, 152, 0.1);
            background: white;
            transform: translateY(-2px);
        }
        
        .form-row { 
            display: grid; 
            grid-template-columns: 1fr 1fr; 
            gap: 25px; 
        }
        
        .form-row.full { 
            grid-template-columns: 1fr; 
        }
        
        .checkbox-group, .radio-group { 
            display: flex; 
            gap: 25px; 
            flex-wrap: wrap; 
        }
        
        .checkbox-item, .radio-item { 
            display: flex; 
            align-items: center; 
            gap: 10px; 
            padding: 8px 15px;
            background: #f8f9fa;
            border-radius: 8px;
            transition: background 0.3s ease;
        }
        
        .checkbox-item:hover, .radio-item:hover {
            background: #eef2f7;
        }
        
        .btn-group { 
            display: flex; 
            gap: 15px; 
            justify-content: flex-end; 
            margin-top: 35px; 
            flex-wrap: wrap; 
        }
        
        .btn { 
            padding: 14px 28px; 
            border: none; 
            border-radius: 10px; 
            font-size: 16px; 
            font-weight: 600; 
            cursor: pointer; 
            transition: all 0.4s cubic-bezier(0.25, 0.46, 0.45, 0.94);
            display: inline-flex;
            align-items: center;
            gap: 8px;
            position: relative;
            overflow: hidden;
        }
        
        .btn::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
            transition: left 0.5s;
        }
        
        .btn:hover::before {
            left: 100%;
        }
        
        .btn-primary { 
            background: linear-gradient(135deg, #2a5298 0%, #1a2f6b 100%); 
            color: white; 
            box-shadow: 0 4px 15px rgba(42, 82, 152, 0.3);
        }
        
        .btn-primary:hover { 
            background: linear-gradient(135deg, #1a2f6b 0%, #2a5298 100%); 
            transform: translateY(-3px); 
            box-shadow: 0 8px 25px rgba(42, 82, 152, 0.4);
        }
        
        .btn-secondary { 
            background: #f8f9fa; 
            color: #2c3e50;
            border: 2px solid #e0e6ed;
        }
        
        .btn-secondary:hover { 
            background: #e9ecef;
            border-color: #2a5298;
            color: #2a5298;
            transform: translateY(-3px);
        }
        
        .btn-danger { 
            background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%); 
            color: white; 
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.3);
        }
        
        .btn-danger:hover { 
            background: linear-gradient(135deg, #c0392b 0%, #e74c3c 100%); 
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(231, 76, 60, 0.4);
        }
        
        .btn-success { 
            background: linear-gradient(135deg, #27ae60 0%, #219653 100%); 
            color: white; 
            box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
        }
        
        .btn-success:hover { 
            background: linear-gradient(135deg, #219653 0%, #27ae60 100%); 
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(39, 174, 96, 0.4);
        }
        
        .alert { 
            padding: 18px 20px; 
            border-radius: 12px; 
            margin-bottom: 25px; 
            display: none; 
            border-left: 5px solid;
            animation: slideInRight 0.5s ease;
            position: relative;
        }
        
        .alert.show { 
            display: block; 
        }
        
        .alert-success { 
            background: #e8f5e8; 
            color: #065f46; 
            border-left-color: #059669;
            border: 1px solid #a3e9c4;
        }
        
        .alert-error { 
            background: #fee2e2; 
            color: #7f1d1d; 
            border-left-color: #dc2626;
            border: 1px solid #fca5a5;
        }
        
        .alert-warning { 
            background: #fef3c7; 
            color: #92400e; 
            border-left-color: #f59e0b;
            border: 1px solid #fcd34d;
        }
        
        .stat-card { 
            background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%); 
            border-radius: 16px; 
            padding: 30px; 
            text-align: center; 
            box-shadow: 0 6px 20px rgba(0,0,0,0.08); 
            border-top: 5px solid #2a5298;
            border: 1px solid #eef2f7;
            transition: transform 0.3s ease;
            position: relative;
            overflow: hidden;
        }
        
        .stat-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 5px;
            background: linear-gradient(135deg, #2a5298 0%, #1a2f6b 100%);
        }
        
        .stat-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 12px 30px rgba(0,0,0,0.12);
        }
        
        .stat-card.closed { border-top-color: #27ae60; }
        .stat-card.pending { border-top-color: #f39c12; }
        .stat-card.within { border-top-color: #27ae60; }
        .stat-card.outof { border-top-color: #e67e22; }
        .stat-card.overdue { border-top-color: #e74c3c; }
        
        .stat-number { 
            font-size: 46px; 
            font-weight: 800; 
            color: #2c3e50; 
            margin: 15px 0; 
            text-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        
        .stat-label { 
            font-size: 14px; 
            color: #64748b; 
            text-transform: uppercase; 
            font-weight: 700; 
            letter-spacing: 0.5px;
        }
        
        .stat-grid { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); 
            gap: 20px; 
            margin-bottom: 35px; 
        }
        
        .status-badge { 
            display: inline-block; 
            padding: 8px 16px; 
            border-radius: 20px; 
            font-size: 12px; 
            font-weight: 700; 
            letter-spacing: 0.5px;
        }
        
        .status-within_time { background: #d1fae5; color: #065f46; border: 1px solid #a3e9c4; }
        .status-out_of_time { background: #fef3c7; color: #92400e; border: 1px solid #fcd34d; }
        .status-overdue { background: #fee2e2; color: #7f1d1d; border: 1px solid #fca5a5; }
        .status-closed { background: #27ae60; color: white; }
        .status-pending { background: #3498db; color: white; }
        
        .option-card { 
            background: white; 
            border-radius: 16px; 
            padding: 35px; 
            text-align: center; 
            cursor: pointer; 
            transition: all 0.4s cubic-bezier(0.25, 0.46, 0.45, 0.94); 
            border: 2px solid #eef2f7; 
            margin-bottom: 25px;
            position: relative;
            overflow: hidden;
        }
        
        .option-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: linear-gradient(135deg, #2a5298 0%, #1a2f6b 100%);
            transform: scaleX(0);
            transition: transform 0.4s ease;
        }
        
        .option-card:hover { 
            transform: translateY(-8px); 
            box-shadow: 0 15px 40px rgba(0,0,0,0.15); 
            border-color: #2a5298;
        }
        
        .option-card:hover::before {
            transform: scaleX(1);
        }
        
        .option-icon { 
            font-size: 56px; 
            margin-bottom: 20px; 
            display: block;
            transition: transform 0.4s ease;
        }
        
        .option-card:hover .option-icon {
            transform: scale(1.1);
        }
        
        .option-card h2 { 
            color: #2c3e50; 
            margin-bottom: 12px; 
            font-size: 22px;
        }
        
        .search-box { 
            margin-bottom: 25px; 
        }
        
        .search-box input { 
            border: 2px solid #e0e6ed; 
            padding: 14px 20px; 
            background: white;
            font-size: 16px;
        }
        
        .dispatch-list { 
            margin-top: 25px; 
        }
        
        .dispatch-item { 
            background: linear-gradient(135deg, #f8fafc 0%, #ffffff 100%); 
            padding: 20px; 
            border-radius: 12px; 
            margin-bottom: 15px; 
            border-left: 5px solid #2a5298;
            border: 1px solid #eef2f7;
            transition: all 0.3s ease;
        }
        
        .dispatch-item:hover {
            transform: translateX(5px);
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }
        
        .fresh-form-indicator { 
            background: linear-gradient(135deg, #e8f5e8 0%, #d1fae5 100%); 
            border: 2px solid #27ae60; 
            padding: 15px 20px; 
            border-radius: 12px; 
            margin-bottom: 25px; 
            text-align: center;
            position: relative;
            overflow: hidden;
        }
        
        .fresh-form-indicator::before {
            content: '🔄';
            position: absolute;
            left: 20px;
            top: 50%;
            transform: translateY(-50%);
            font-size: 20px;
        }
        
        .dispatch-status { 
            display: inline-block; 
            padding: 6px 14px; 
            border-radius: 15px; 
            font-size: 12px; 
            font-weight: 700; 
        }
        
        .status-processed { 
            background: #fff3cd; 
            color: #856404; 
            border: 1px solid #ffeaa7; 
        }
        
        .status-completed { 
            background: #d1fae5; 
            color: #065f46; 
            border: 1px solid #a3e9c4; 
        }
        
        .section-title {
            font-size: 28px;
            font-weight: 700;
            color: #2c3e50;
            margin-bottom: 10px;
            position: relative;
            padding-bottom: 15px;
        }
        
        .section-title::after {
            content: '';
            position: absolute;
            bottom: 0;
            left: 0;
            width: 60px;
            height: 4px;
            background: linear-gradient(135deg, #2a5298 0%, #1a2f6b 100%);
            border-radius: 2px;
        }
        
        .section-subtitle {
            font-size: 16px;
            color: #64748b;
            margin-bottom: 30px;
            font-weight: 500;
        }
        
        .application-list {
            margin: 20px 0;
        }
        
        .application-item {
            background: #f8f9fa;
            border: 1px solid #e9ecef;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 10px;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        
        .application-item:hover {
            background: #e9ecef;
            border-color: #2a5298;
        }
        
        .application-item.selected {
            background: #d1ecf1;
            border-color: #17a2b8;
        }
        
        .application-details {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .application-info {
            flex: 1;
        }
        
        .application-actions {
            margin-left: 15px;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        @keyframes slideInRight {
            from { opacity: 0; transform: translateX(30px); }
            to { opacity: 1; transform: translateX(0); }
        }
        
        @media (max-width: 768px) {
            .form-row { grid-template-columns: 1fr; }
            .stat-grid { grid-template-columns: repeat(2, 1fr); }
            .btn-group { flex-direction: column; }
            .btn { width: 100%; justify-content: center; }
            header { flex-direction: column; text-align: center; padding: 20px; }
            .nav-tabs { flex-wrap: wrap; padding: 0 15px; gap: 10px; }
            .nav-tab { padding: 15px 20px; flex: 1; min-width: 120px; justify-content: center; }
            .container { margin: 25px auto; padding: 0 15px; }
            .card { padding: 25px; margin-bottom: 20px; }
            header img { width: 70px; height: 70px; }
            header h1 { font-size: 22px; }
            .section-title { font-size: 24px; }
        }
        
        /* Loading animation */
        .loading {
            display: inline-block;
            width: 20px;
            height: 20px;
            border: 3px solid #f3f3f3;
            border-top: 3px solid #2a5298;
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin-right: 10px;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
    </style>
</head>
<body>
    <header>
        <img src="static/sku_logo.png" alt="SKU Logo">
        <div class="header-content">
            <h1>SRI KRISHNADEVARAYA UNIVERSITY<br>ANANTHAPURAMU</h1>
            <p>TAPAL SECTION - DATA ENTRY MANAGEMENT SYSTEM</p>
        </div>
    </header>
    
    <div class="nav-tabs">
        <button class="nav-tab active" onclick="switchTab('inward')">📥 IN-WARD</button>
        <button class="nav-tab" onclick="switchTab('outward')">📦 OUT-WARD</button>
        <button class="nav-tab" onclick="switchTab('admin')">👤 ADMIN</button>
    </div>
    
    <div class="container">
        <!-- IN-WARD SECTION -->
        <div id="inward" class="section active">
            <div id="applicationSelection" class="card">
                <h2 class="section-title">Application Type Selection</h2>
                <p class="section-subtitle">Choose the appropriate application type for processing</p>
                
                <div class="form-row full">
                    <div class="option-card" onclick="selectFormType('regular')">
                        <div class="option-icon">📋</div>
                        <h2>Regular Application</h2>
                        <p style="color: #7f8c8d;">For UG & PG certificate processing</p>
                    </div>
                </div>
                <div class="form-row full">
                    <div class="option-card" onclick="selectFormType('distance')">
                        <div class="option-icon">🌐</div>
                        <h2>Distance Application</h2>
                        <p style="color: #7f8c8d;">For distance education programs</p>
                    </div>
                </div>
                <div class="form-row full">
                    <div class="option-card" onclick="selectFormType('confidential')">
                        <div class="option-icon">🔐</div>
                        <h2>Confidential Application</h2>
                        <p style="color: #7f8c8d;">For confidential and sensitive items</p>
                    </div>
                </div>
            </div>
            
            <!-- REGULAR FORM -->
            <div id="regularForm" class="card" style="display: none;">
                <div class="fresh-form-indicator">
                    
                </div>
                <h2 class="section-title">Regular Application Form</h2>
                <p class="section-subtitle">Complete all required fields for regular certificate processing</p>
                
                <div id="regularAlert" class="alert"></div>
                
                <div class="form-group required">
                    <label>Full Name:</label>
                    <input type="text" id="regName" maxlength="100" placeholder="Enter applicant's full name">
                </div>
                
                <div class="form-row">
                    <div class="form-group required">
                        <label>Register Number:</label>
                        <input type="text" id="regRegister" placeholder="Enter university register number">
                    </div>
                    <div class="form-group required">
                        <label>Hall Ticket Number:</label>
                        <input type="text" id="regHallTicket" placeholder="Enter examination hall ticket number">
                    </div>
                </div>
                
                <div class="form-group">
                    <label>Address:</label>
                    <textarea id="regAddress" maxlength="500" style="min-height: 80px;" placeholder="Enter communication address (optional)"></textarea>
                </div>
                
                <div class="form-row">
                    <div class="form-group required">
                        <label>Degree Level:</label>
                        <select id="regDegree" onchange="updateRegularForm()">
                            <option value="">Select Degree Level</option>
                            <option value="ug">UG - Under Graduate</option>
                            <option value="pg">PG - Post Graduate</option>
                            <option value="mphil">Master of Philosophy</option>
                        </select>
                    </div>
                    <div class="form-group required">
                        <label>Course Program:</label>
                        <select id="regCourse" onchange="updatePGTableNumber()">
                            <option value="">Select Course Program</option>
                        </select>
                    </div>
                </div>
                
                <!-- NEW: Table Number Selection for UG -->
                <div id="regTableGroup" class="form-group required" style="display: none;">
                    <label>Table Number:</label>
                    <select id="regTableNumber" onchange="updateCollegesByTable()">
                        <option value="">Select Table Number</option>
                        <option value="B-1">B-1</option>
                        <option value="B-2">B-2</option>
                        <option value="B-3">B-3</option>
                        <option value="B-4">B-4</option>
                        <option value="B-5">B-5</option>
                        <option value="B-6">B-6</option>
                        <option value="B-7">B-7</option>
                        <option value="B-8">B-8</option>
                        <option value="B-9">B-9</option>
                        <option value="B-10">B-10</option>
                        <option value="B-11">B-11</option>
                        <option value="B-12">B-12</option>
                    </select>
                </div>
                
                <!-- College Selection (only for UG) -->
                <div id="regCollegeGroup" class="form-group required" style="display: none;">
                    <label>College Institution:</label>
                    <select id="regCollege">
                        <option value="">Select College</option>
                    </select>
                </div>
                
                <!-- PG College Code Input with Auto Table Number -->
                <div id="regPGCollegeGroup" class="form-group required" style="display: none;">
                    <div class="form-row">
                        <div class="form-group required">
                            <label>College Code (PG/MPhil):</label>
                            <input type="text" id="regPGCollege" placeholder="Enter college code for PG/MPhil programs">
                        </div>
                        <div class="form-group">
                            <label>Table Number (Auto):</label>
                            <input type="text" id="regPGTableNumber" readonly style="background: #f8f9fa;">
                        </div>
                    </div>
                </div>
                
                <div class="form-group required">
                    <label>Certificate Type:</label>
                    <select id="regCertificateType" onchange="updateRequiredDocs('regular')">
                        <option value="">Select Certificate Type</option>
                        <option value="Provisional Certificate">Provisional Certificate</option>
                        <option value="Migration Certificate">Migration Certificate</option>
                        <option value="Convocation Certificate">Convocation Certificate</option>
                        <option value="Transcript">Transcript</option>
                        <option value="Genuineness Certificate">Genuineness Certificate</option>
                        <option value="Duplicate Marks Memo">Duplicate Marks Memo</option>
                        <option value="Duplicate Certificate">Duplicate Certificate</option>
                        <option value="Mphill/Phd Provisional certificate">Mphill/Phd Provisional certificate</option>
                        <option value="Mphill/Phd Convocation certificate">Mphill/Phd Convocation certificate</option>
                    </select>
                </div>
                
                <div class="form-group required">
                    <label>Document Mode:</label>
                    <div class="radio-group">
                        <div class="radio-item">
                            <input type="radio" name="regMode" value="Original" checked>
                            <label style="margin-bottom: 0;">Original Document</label>
                        </div>
                        <div class="radio-item">
                            <input type="radio" name="regMode" value="Duplicate">
                            <label style="margin-bottom: 0;">Duplicate Document</label>
                        </div>
                    </div>
                </div>
                
                <div id="regRequiredDocs" class="form-group">
                    <label>Required Documents Checklist:</label>
                    <div class="checkbox-group" id="regDocsCheckbox"></div>
                </div>
                
                <div class="form-group required">
                    <label>Payment Method:</label>
                    <select id="regPaymentMode" onchange="updatePaymentFields('regular')">
                        <option value="">Select Payment Method</option>
                        <option value="Online">Online Payment (SBI)</option>
                        <option value="Offline">Offline Payment (SBI)</option>
                        <option value="DD">Demand Draft (Manual Bank)</option>
                    </select>
                </div>
                
                <div id="regPaymentFields"></div>
                
                <div class="btn-group">
                    <button class="btn btn-secondary" onclick="goBack()">
                        ← Back to Selection
                    </button>
                    <button class="btn btn-secondary" onclick="resetCurrentForm('regular')">
                        🔄 Reset Form
                    </button>
                    <button class="btn btn-primary" onclick="submitRegularForm()">
                        ✓ Save Application
                    </button>
                </div>
            </div>
            
            <!-- DISTANCE FORM -->
            <div id="distanceForm" class="card" style="display: none;">
                <div class="fresh-form-indicator">
                    
                </div>
                <h2 class="section-title">Distance Education Application</h2>
                <p class="section-subtitle">Complete all required fields for distance education programs</p>
                
                <div id="distanceAlert" class="alert"></div>
                
                <div class="form-group required">
                    <label>Full Name:</label>
                    <input type="text" id="distName" maxlength="100" placeholder="Enter applicant's full name">
                </div>
                
                <div class="form-row">
                    <div class="form-group required">
                        <label>Register Number:</label>
                        <input type="text" id="distRegister" placeholder="Enter university register number">
                    </div>
                    <div class="form-group required">
                        <label>Hall Ticket Number:</label>
                        <input type="text" id="distHallTicket" placeholder="Enter examination hall ticket number">
                    </div>
                </div>
                
                <div class="form-group">
                    <label>Address:</label>
                    <textarea id="distAddress" maxlength="500" style="min-height: 80px;" placeholder="Enter communication address (optional)"></textarea>
                </div>
                
                <div class="form-row">
                    <div class="form-group required">
                        <label>Degree Level:</label>
                        <select id="distDegree" onchange="updateDistanceCourses()">
                            <option value="">Select Degree Level</option>
                            <option value="ug">UG - Under Graduate</option>
                            <option value="pg">PG - Post Graduate</option>
                            <option value="mphil">Master of Philosophy</option>
                            <option value="bed_ug">B.Ed.(UG) - Bachelor of Education (Undergraduate)</option>
                            <option value="bed_pg">B.Ed.(PG) - Bachelor of Education (Postgraduate)</option>
                        </select>
                    </div>
                    <div class="form-group required">
                        <label>Course Program:</label>
                        <select id="distCourse">
                            <option value="">Select Course Program</option>
                        </select>
                    </div>
                </div>
                
                <!-- Distance education has fixed table number D-1 -->
                <div class="form-group">
                    <label>Table Number:</label>
                    <input type="text" id="distTableNumber" readonly value="D-1" style="background: #f8f9fa; font-weight: bold; font-size: 16px; color: #1a1a1a; border: 2px solid #3498db; padding: 10px;">
                </div>
                                <div class="form-group">
                    <label>College Code (Optional):</label>
                    <input type="text" id="distCollegeCode" placeholder="Enter college code (optional)" maxlength="50">
                </div>
                
                <div class="form-group required">
                    <label>Certificate Type:</label>
                    <select id="distCertificateType" onchange="updateRequiredDocs('distance')">
                        <option value="">Select Certificate Type</option>
                        <option value="Provisional Certificate">Provisional Certificate</option>
                        <option value="Migration Certificate">Migration Certificate</option>
                        <option value="Convocation Certificate">Convocation Certificate</option>
                        <option value="Transcript">Transcript</option>
                        <option value="Genuineness Certificate">Genuineness Certificate</option>
                        <option value="Duplicate Marks Memo">Duplicate Marks Memo</option>
                        <option value="Duplicate Certificate">Duplicate Certificate</option>
                        <option value="Mphill/Phd Provisional certificate">Mphill/Phd Provisional certificate</option>
                        <option value="Mphill/Phd Convocation certificate">Mphill/Phd Convocation certificate</option>
                    </select>
                </div>
                
                <div class="form-group required">
                    <label>Document Mode:</label>
                    <div class="radio-group">
                        <div class="radio-item">
                            <input type="radio" name="distMode" value="Original" checked>
                            <label style="margin-bottom: 0;">Original Document</label>
                        </div>
                        <div class="radio-item">
                            <input type="radio" name="distMode" value="Duplicate">
                            <label style="margin-bottom: 0;">Duplicate Document</label>
                        </div>
                    </div>
                </div>
                
                <div id="distRequiredDocs" class="form-group">
                    <label>Required Documents Checklist:</label>
                    <div class="checkbox-group" id="distDocsCheckbox"></div>
                </div>
                
                <div class="form-group required">
                    <label>Payment Method:</label>
                    <select id="distPaymentMode" onchange="updatePaymentFields('distance')">
                        <option value="">Select Payment Method</option>
                        <option value="Online">Online Payment (SBI)</option>
                        <option value="Offline">Offline Payment (SBI)</option>
                        <option value="DD">Demand Draft (Manual Bank)</option>
                    </select>
                </div>
                
                <div id="distPaymentFields"></div>
                
                <div class="btn-group">
                    <button class="btn btn-secondary" onclick="goBack()">
                        ← Back to Selection
                    </button>
                    <button class="btn btn-secondary" onclick="resetCurrentForm('distance')">
                        🔄 Reset Form
                    </button>
                    <button class="btn btn-primary" onclick="submitDistanceForm()">
                        ✓ Save Application
                    </button>
                </div>
            </div>
            
            <!-- CONFIDENTIAL FORM -->
            <div id="confidentialForm" class="card" style="display: none;">
                <div class="fresh-form-indicator">
                    
                </div>
                <h2 class="section-title">Confidential Application</h2>
                <p class="section-subtitle">For confidential and sensitive document processing</p>
                
                <div id="confAlert" class="alert"></div>
                
                <div class="form-group required">
                    <label>College Code:</label>
                    <input type="text" id="confCollege" placeholder="Enter college institution code">
                </div>
                
                <div class="form-group required">
                    <label>Hall Ticket Number:</label>
                    <input type="text" id="confHallTicket" placeholder="Enter examination hall ticket number">
                </div>
                
                <div class="form-group required">
                    <label>Application Details:</label>
                    <textarea id="confDetails" maxlength="500" style="min-height: 120px;" placeholder="Provide detailed description of confidential application"></textarea>
                </div>
                
                <!-- Confidential application has fixed table number C-1 -->
                <div class="form-group">
                    <label>Table Number:</label>
                    <input type="text" id="confTableNumber" readonly value="C-1" style="background: #f8f9fa; font-weight: bold; font-size: 16px; color: #1a1a1a; border: 2px solid #3498db; padding: 10px;">
                </div>
                
                <div class="form-group required">
                    <label>Payment Method:</label>
                    <select id="confPaymentMode" onchange="updatePaymentFields('confidential')">
                        <option value="">Select Payment Method</option>
                        <option value="Online">Online Payment (SBI)</option>
                        <option value="Offline">Offline Payment (SBI)</option>
                        <option value="DD">Demand Draft (Manual Bank)</option>
                    </select>
                </div>
                
                <div id="confPaymentFields"></div>
                
                <div class="btn-group">
                    <button class="btn btn-secondary" onclick="goBack()">
                        ← Back to Selection
                    </button>
                    <button class="btn btn-secondary" onclick="resetCurrentForm('confidential')">
                        🔄 Reset Form
                    </button>
                    <button class="btn btn-primary" onclick="submitConfidentialForm()">
                        ✓ Save Application
                    </button>
                </div>
            </div>
        </div>
        
        <!-- OUT-WARD SECTION -->
        <div id="outward" class="section">
            <div class="card">
                <h2 class="section-title">Dispatch Management System</h2>
                <p class="section-subtitle">Manage outgoing applications and track dispatch status</p>
                
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 25px; margin-bottom: 35px;">
                    <div class="option-card" style="border-color: #3498db; padding: 40px;" onclick="showDispatchSection('processing')">
                        <div style="font-size: 56px; margin-bottom: 15px;">📤</div>
                        <h3 style="color: #2c3e50; margin-bottom: 10px;">Dispatch Processing</h3>
                        <p style="color: #7f8c8d;">Process outgoing applications for dispatch</p>
                    </div>
                    <div class="option-card" style="border-color: #27ae60; padding: 40px;" onclick="showDispatchSection('closure')">
                        <div style="font-size: 56px; margin-bottom: 15px;">✓</div>
                        <h3 style="color: #2c3e50; margin-bottom: 10px;">Dispatch Closure</h3>
                        <p style="color: #7f8c8d;">Close processed dispatches with tracking</p>
                    </div>
                </div>
                
                <!-- DISPATCH PROCESSING -->
                <div id="dispatchProcessing" style="display: none;">
                    <h3 style="margin-bottom: 20px; color: #2c3e50;">📤 Dispatch Processing</h3>
                    <p style="margin-bottom: 20px; color: #7f8c8d;">Process applications for dispatch by verifying hall ticket and selecting dispatch mode</p>
                    
                    <div class="form-row">
                        <div class="form-group required">
                            <label>Hall Ticket Number:</label>
                            <input type="text" id="dispatchHallTicket" placeholder="Enter hall ticket number for verification">
                        </div>
                        <div class="form-group" style="display: flex; align-items: flex-end;">
                            <button class="btn btn-primary" style="width: 100%;" onclick="verifyDispatchHallTicket()">
                                🔍 Verify & Proceed
                            </button>
                        </div>
                    </div>
                    
                    <div id="pendingApplicationsList" class="application-list" style="display: none;">
                        <h4 style="margin-bottom: 15px; color: #2c3e50;">📋 Pending Applications</h4>
                        <p style="margin-bottom: 15px; color: #7f8c8d;">Select an application to process for dispatch:</p>
                        <div id="pendingApplications"></div>
                    </div>
                    
                    <div id="dispatchDetails" style="display: none; background: #f8fafc; padding: 25px; border-radius: 12px; margin-bottom: 25px; border-left: 5px solid #3498db; border: 1px solid #eef2f7;"></div>
                    
                    <div class="form-group required" id="dispatchModeGroup" style="display: none;">
                        <label>Dispatch Mode:</label>
                        <select id="dispatchMode">
                            <option value="">Select Dispatch Method</option>
                            <option value="POST">📮 Postal Service</option>
                            <option value="COURIER">🚚 Courier Service</option>
                            <option value="BY HAND">✋ Hand Delivery</option>
                        </select>
                    </div>
                    
                    <div class="btn-group" id="dispatchSaveGroup" style="display: none;">
                        <button class="btn btn-secondary" onclick="goBackDispatch()">
                            ← Back to Options
                        </button>
                        <button class="btn btn-success" onclick="saveDispatch()">
                            ✓ Process Dispatch
                        </button>
                    </div>
                </div>
                
                <!-- DISPATCH CLOSURE -->
                <div id="dispatchClosure" style="display: none;">
                    <h3 style="margin-bottom: 20px; color: #2c3e50;">✓ Dispatch Closure</h3>
                    <p style="margin-bottom: 20px; color: #7f8c8d;">Close processed dispatches by providing tracking ID or reason for closure</p>
                    
                    <div class="form-row">
                        <div class="form-group required">
                            <label>Hall Ticket Number:</label>
                            <input type="text" id="closureHallTicket" placeholder="Enter hall ticket number for closure">
                        </div>
                        <div class="form-group" style="display: flex; align-items: flex-end;">
                            <button class="btn btn-primary" style="width: 100%;" onclick="verifyClosureHallTicket()">
                                🔍 Verify Application
                            </button>
                        </div>
                    </div>
                    
                    <div id="processedDispatchesList" class="application-list" style="display: none;">
                        <h4 style="margin-bottom: 15px; color: #2c3e50;">📦 Processed Dispatches</h4>
                        <p style="margin-bottom: 15px; color: #7f8c8d;">Select a dispatch to close:</p>
                        <div id="processedDispatches"></div>
                    </div>
                    
                    <div id="closureDetails" style="display: none; background: #f8fafc; padding: 25px; border-radius: 12px; margin-bottom: 25px; border-left: 5px solid #27ae60; border: 1px solid #eef2f7;"></div>
                    
                    <div id="closureFields" style="display: none;">
                        <div class="form-group required">
                            <label>Tracking ID / Closure Reason:</label>
                            <input type="text" id="closureTrackingId" placeholder="Enter tracking number or reason for dispatch closure">
                        </div>
                        
                        <div class="btn-group">
                            <button class="btn btn-secondary" onclick="goBackDispatch()">
                                ← Back to Options
                            </button>
                            <button class="btn btn-success" onclick="closeDispatch()">
                                ✓ Confirm Closure
                            </button>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- ADMIN SECTION -->
        <div id="admin" class="section">
            <div id="adminLogin" class="card" style="max-width: 450px; margin: 60px auto;">
                <h2 style="text-align: center; margin-bottom: 35px; color: #2c3e50;">🔐 Administrator Login</h2>
                <div id="loginAlert" class="alert"></div>
                
                <div class="form-group required">
                    <label>Username:</label>
                    <input type="text" id="adminUsername" placeholder="Enter administrator username">
                </div>
                
                <div class="form-group required">
                    <label>Password:</label>
                    <input type="password" id="adminPassword" placeholder="Enter administrator password">
                </div>
                
                <div class="btn-group">
                    <button class="btn btn-primary" onclick="adminLogin()" style="width: 100%;">
                        🔑 Login to Dashboard
                    </button>
                </div>
            </div>
            
            <div id="adminDashboard" style="display: none;">
                <div class="card">
                    <h2 class="section-title">Administrative Dashboard</h2>
                    <p class="section-subtitle">Comprehensive overview of application statistics and system reports</p>
                    
                    <h3 style="margin-bottom: 25px; color: #2c3e50;">📊 Application Statistics</h3>
                    <div class="stat-grid">
                        <div class="stat-card">
                            <div class="stat-label">📅 Today's Applications</div>
                            <div class="stat-number" id="statToday">0</div>
                        </div>
                        <div class="stat-card closed">
                            <div class="stat-label">✓ Closed Applications</div>
                            <div class="stat-number" id="statClosed">0</div>
                        </div>
                        <div class="stat-card pending">
                            <div class="stat-label">⏳ Pending Applications</div>
                            <div class="stat-number" id="statPending">0</div>
                        </div>
                        <div class="stat-card within">
                            <div class="stat-label">✅ Within Timeframe</div>
                            <div class="stat-number" id="statWithin">0</div>
                        </div>
                        <div class="stat-card outof">
                            <div class="stat-label">⚠️ Approaching Deadline</div>
                            <div class="stat-number" id="statOutOf">0</div>
                        </div>
                        <div class="stat-card overdue">
                            <div class="stat-label">🔴 Overdue Applications</div>
                            <div class="stat-number" id="statOverdue">0</div>
                        </div>
                    </div>
                </div>
                
                <div class="card">
                    <h3 style="margin-bottom: 25px; color: #2c3e50;">📥 Report Generation</h3>
                    
                    <div style="margin-bottom: 30px;">
                        <h4 style="margin-bottom: 18px; color: #2c3e50;">📅 Date Range Report</h4>
                        <div class="form-row">
                            <div class="form-group">
                                <label>Start Date:</label>
                                <input type="date" id="reportFromDate">
                            </div>
                            <div class="form-group">
                                <label>End Date:</label>
                                <input type="date" id="reportToDate">
                            </div>
                        </div>
                        <button class="btn btn-primary" onclick="exportDateRangeReport()">
                            📥 Download Date Range Report
                        </button>
                    </div>
                    
                    <hr style="margin: 30px 0; border: 1px solid #eef2f7;">
                    
                    <div>
                        <h4 style="margin-bottom: 18px; color: #2c3e50;">🏫 Table-Wise Report</h4>
                        <div class="form-row">
                            <div class="form-group">
                                <label>Select Table:</label>
                                <select id="reportTableNumber">
                                    <option value="">Select Table Category</option>
                                    <option value="B-1">B-1</option><option value="B-2">B-2</option><option value="B-3">B-3</option><option value="B-4">B-4</option><option value="B-5">B-5</option><option value="B-6">B-6</option>
                                    <option value="B-7">B-7</option><option value="B-8">B-8</option><option value="B-9">B-9</option><option value="B-10">B-10</option>
                                    <option value="B-11">B-11</option><option value="B-12">B-12</option>
                                    <option value="M-1">M-1 (M.A)</option><option value="M-2">M-2 (M.Sc)</option><option value="M-3">M-3 (M.Com)</option><option value="M-4">M-4 (M.C.A)</option><option value="M-5">M-5 (M.B.A)</option><option value="M-6">M-6 (M.Pharmacy)</option>
                                    <option value="L-1">L-1 (LLB)</option><option value="G-1">G-1 (B.Tech/M.Tech/PhD)</option><option value="D-1">D-1 (Distance)</option><option value="C-1">C-1 (Confidential)</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Start Date:</label>
                                <input type="date" id="reportTableFromDate">
                            </div>
                            <div class="form-group">
                                <label>End Date:</label>
                                <input type="date" id="reportTableToDate">
                            </div>
                        </div>
                        <button class="btn btn-primary" onclick="exportTableWiseReport()">
                            📥 Download Table Report
                        </button>
                    </div>
                    
                    
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let allColleges = [];
        let isAdminLoggedIn = false;
        let autoRefreshInterval = null;
        let selectedAppId = null;
        let selectedDispatchAppId = null;
        
        const courseMappings = {
            ug: ['BA', 'B.Com', 'B.Sc', 'B.C.A', 'B.Pharmacy', 'B.Ed'],
            pg: ['M.A', 'M.Sc', 'M.Com', 'M.C.A', 'M.B.A', 'M.Pharmacy', 'LLB', 'BTech/MTech'],
            mphil: ['Ph.D'],
            bed_ug: ['B.Ed.(Physical Science)', 'B.Ed.(Biological Science)', 'B.Ed.(Mathematics)', 'B.Ed.(Social Studies)', 'B.Ed.(English)'],
            bed_pg: ['B.Ed.(Physical Science)', 'B.Ed.(Biological Science)', 'B.Ed.(Mathematics)', 'B.Ed.(Social Studies)', 'B.Ed.(English)']
        };
        
        const pgCourseTableMap = {
            'M.A': 'M-1', 'M.Sc': 'M-2', 'M.Com': 'M-3', 'M.C.A': 'M-4',
            'M.B.A': 'M-5', 'M.Pharmacy': 'M-6', 'LLB': 'L-1', 'BTech/MTech': 'G-1', 'Ph.D': 'G-1'
        };
        
        // UPDATED REQUIRED DOCS WITH NEW FIELDS
        const requiredDocs = {
            'Duplicate Marks Memo': ['SBI Bank Challan', 'Application Form', 'Marks Memos', 'Cloth Cover', 'Postal Stamp', 'Aadhar Card'],
            'Provisional Certificate': ['SBI Bank Challan', 'Application Form', 'Marks Memos', 'Convocation Certificate', 'Migration Certificate', 'Intermediate Marks', '10th Marks', 'Aadhar Card', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
            'Convocation Certificate': ['SBI Bank Challan', 'Application Form', 'CCM & Provisional', 'Convocation Certificate', 'Intermediate Marks', '10th Marks', 'Aadhar Card', 'Cloth Cover', 'Photos', 'Gazetted Signature', 'APPAR ID'],
            'Transcript': ['SBI Bank Challan', 'Application Form', 'CCM & Provisional', 'Certificates', 'Cloth Cover', 'Postal Stamp', 'Aadhar Card'],
            'Migration Certificate': ['SBI Bank Challan', 'Application Form', 'Degree Certificate', 'Marks Memos', 'Provisional Certificate', 'Aadhar Card', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
            'Genuineness Certificate': ['Notification', 'Proceeding Copy', 'Journals', 'Marks Memos', 'Convocation Certificate', 'Aadhar Card', 'Photos', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
            'Duplicate Certificate': ['SBI Bank Challan', 'Application Form', 'Original Certificate', 'Aadhar Card', 'Photos', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
            'Mphill/Phd Provisional certificate': ['Notification', 'Proceeding Copy', 'Journals', 'Marks Memos', 'Convocation Certificate', 'Aadhar Card', 'Photos', 'Cloth Cover', 'Postal Stamp', 'APPAR ID'],
            'Mphill/Phd Convocation certificate': ['Notification', 'Proceeding Copy', 'Journals', 'Marks Memos', 'MPhil Provisional', 'Convocation Certificate', 'Aadhar Card', 'Photos', 'Gazetted Signature', 'Cloth Cover', 'Postal Stamp', 'APPAR ID']
        };
        
        document.addEventListener('DOMContentLoaded', function() {
            loadColleges();
            resetAllForms();
        });
        
        function loadColleges() {
            fetch('/api/colleges')
                .then(r => r.json())
                .then(data => {
                    allColleges = data;
                })
                .catch(err => {
                    console.error('Error loading colleges:', err);
                });
        }
        
        function switchTab(tab) {
            document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
            document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
            
            const targetSection = document.getElementById(tab);
            const targetTab = event.target;
            
            if (targetSection) targetSection.classList.add('active');
            if (targetTab) targetTab.classList.add('active');
            
            if (tab === 'admin' && isAdminLoggedIn) {
                loadAdminData();
                if (autoRefreshInterval) clearInterval(autoRefreshInterval);
                autoRefreshInterval = setInterval(loadAdminData, 10000);
            } else if (tab === 'inward') {
                resetAllForms();
            }
        }
        
        function selectFormType(type) {
            document.getElementById('applicationSelection').style.display = 'none';
            resetAllForms();
            
            const formElement = document.getElementById(type + 'Form');
            if (formElement) {
                formElement.style.display = 'block';
            }
            
            // Initialize specific form type
            if (type === 'regular') {
                updateRegularForm();
            } else if (type === 'distance') {
                updateDistanceCourses();
            }
        }
        
        function resetAllForms() {
            resetForm('regular');
            resetForm('distance');
            resetForm('confidential');
        }
        
        function resetForm(formType) {
            const form = document.getElementById(formType + 'Form');
            if (!form) return;
            
            const inputs = form.querySelectorAll('input, select, textarea');
            inputs.forEach(input => {
                if (input.type === 'text' || input.type === 'password' || input.tagName === 'TEXTAREA') {
                    input.value = '';
                } else if (input.type === 'select-one') {
                    input.selectedIndex = 0;
                } else if (input.type === 'radio' || input.type === 'checkbox') {
                    input.checked = false;
                }
            });
            
            // Reset specific fields
            if (formType === 'regular') {
                document.querySelector('input[name="regMode"][value="Original"]').checked = true;
                document.getElementById('regPaymentFields').innerHTML = '';
                document.getElementById('regDocsCheckbox').innerHTML = '';
                document.getElementById('regTableGroup').style.display = 'none';
                document.getElementById('regCollegeGroup').style.display = 'none';
                document.getElementById('regPGCollegeGroup').style.display = 'none';
                document.getElementById('regPGTableNumber').value = '';
            } else if (formType === 'distance') {
                document.getElementById('distTableNumber').value = 'D-1';
                document.getElementById('distCollegeCode').value = '';
                document.querySelector('input[name="distMode"][value="Original"]').checked = true;
                document.getElementById('distPaymentFields').innerHTML = '';
                document.getElementById('distDocsCheckbox').innerHTML = '';
            } else if (formType === 'confidential') {
                document.getElementById('confTableNumber').value = 'C-1';
                document.getElementById('confPaymentFields').innerHTML = '';
            }
        }
        
        function resetCurrentForm(formType) {
            resetForm(formType);
            showAlert(formType + 'Alert', '🔄 Form has been reset to fresh state', 'success');
        }
        
        function goBack() {
            document.getElementById('regularForm').style.display = 'none';
            document.getElementById('distanceForm').style.display = 'none';
            document.getElementById('confidentialForm').style.display = 'none';
            document.getElementById('applicationSelection').style.display = 'block';
            resetAllForms();
        }
        
        function goBackDispatch() {
            document.getElementById('dispatchProcessing').style.display = 'none';
            document.getElementById('dispatchClosure').style.display = 'none';
            resetDispatchForm();
        }
        
        function resetDispatchForm() {
            const dispatchHallTicket = document.getElementById('dispatchHallTicket');
            const dispatchMode = document.getElementById('dispatchMode');
            const closureHallTicket = document.getElementById('closureHallTicket');
            const closureTrackingId = document.getElementById('closureTrackingId');
            
            if (dispatchHallTicket) dispatchHallTicket.value = '';
            if (dispatchMode) dispatchMode.value = '';
            if (closureHallTicket) closureHallTicket.value = '';
            if (closureTrackingId) closureTrackingId.value = '';
            
            document.getElementById('dispatchDetails').style.display = 'none';
            document.getElementById('dispatchModeGroup').style.display = 'none';
            document.getElementById('dispatchSaveGroup').style.display = 'none';
            document.getElementById('closureDetails').style.display = 'none';
            document.getElementById('closureFields').style.display = 'none';
            document.getElementById('pendingApplicationsList').style.display = 'none';
            document.getElementById('processedDispatchesList').style.display = 'none';
            
            selectedAppId = null;
            selectedDispatchAppId = null;
        }
        
        // NEW: Update regular form based on degree selection
        function updateRegularForm() {
            const degree = document.getElementById('regDegree');
            const courseSelect = document.getElementById('regCourse');
            const tableGroup = document.getElementById('regTableGroup');
            const collegeGroup = document.getElementById('regCollegeGroup');
            const pgCollegeGroup = document.getElementById('regPGCollegeGroup');
            
            if (!degree || !courseSelect) return;
            
            courseSelect.innerHTML = '<option value="">Select Course</option>';
            tableGroup.style.display = 'none';
            collegeGroup.style.display = 'none';
            pgCollegeGroup.style.display = 'none';
            document.getElementById('regPGTableNumber').value = '';
            
            if (degree.value === 'ug') {
                tableGroup.style.display = 'block';
                courseMappings.ug.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            } else if (degree.value === 'pg' || degree.value === 'mphil') {
                pgCollegeGroup.style.display = 'block';
                const courses = degree.value === 'pg' ? courseMappings.pg : courseMappings.mphil;
                courses.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            }
        }
        
        // NEW: Update PG table number based on course selection
        function updatePGTableNumber() {
            const course = document.getElementById('regCourse');
            const pgTableNumber = document.getElementById('regPGTableNumber');
            
            if (!course || !pgTableNumber) return;
            
            if (course.value && pgCourseTableMap[course.value]) {
                pgTableNumber.value = pgCourseTableMap[course.value];
            } else {
                pgTableNumber.value = '';
            }
        }
        
        // NEW: Update colleges based on selected table number
        function updateCollegesByTable() {
            const tableNumber = document.getElementById('regTableNumber');
            const collegeSelect = document.getElementById('regCollege');
            const collegeGroup = document.getElementById('regCollegeGroup');
            
            if (!tableNumber || !collegeSelect) return;
            
            collegeSelect.innerHTML = '<option value="">Select College</option>';
            collegeGroup.style.display = 'none';
            
            if (tableNumber.value) {
                collegeGroup.style.display = 'block';
                // Filter colleges by selected table number
                const filteredColleges = allColleges.filter(college => college.table === tableNumber.value);
                
                filteredColleges.forEach(college => {
                    const option = document.createElement('option');
                    option.value = college.code;
                    option.textContent = `${college.code} - ${college.name}`;
                    collegeSelect.appendChild(option);
                });
            }
        }
        
        function updateDistanceCourses() {
            const degree = document.getElementById('distDegree');
            const courseSelect = document.getElementById('distCourse');
            
            if (!degree || !courseSelect) return;
            
            courseSelect.innerHTML = '<option value="">Select Course</option>';
            
            if (degree.value === 'ug') {
                courseMappings.ug.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            } else if (degree.value === 'bed_ug') {
                courseMappings.bed_ug.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            } else if (degree.value === 'bed_pg') {
                courseMappings.bed_pg.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            } else if (degree.value === 'pg') {
                courseMappings.pg.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            } else if (degree.value === 'mphil') {
                courseMappings.mphil.forEach(c => {
                    courseSelect.innerHTML += `<option value="${c}">${c}</option>`;
                });
            }
        }
        
        function updatePaymentFields(formType) {
            let mode = '';
            let fieldsDiv = '';
            
            if (formType === 'regular') {
                mode = document.getElementById('regPaymentMode')?.value || '';
                fieldsDiv = 'regPaymentFields';
            } else if (formType === 'distance') {
                mode = document.getElementById('distPaymentMode')?.value || '';
                fieldsDiv = 'distPaymentFields';
            } else if (formType === 'confidential') {
                mode = document.getElementById('confPaymentMode')?.value || '';
                fieldsDiv = 'confPaymentFields';
            }
            
            const fieldsElement = document.getElementById(fieldsDiv);
            if (!fieldsElement) return;
            
            let html = '';
            if (mode === 'Online' || mode === 'Offline') {
                const label = mode === 'Online' ? 'UTR/Reference Number' : 'Challan Number';
                html = `
                    <div class="form-group required">
                        <label>${label}:</label>
                        <input type="text" class="paymentRef" placeholder="Enter ${label}">
                    </div>
                    <div class="form-row">
                        <div class="form-group required">
                            <label>Amount (₹):</label>
                            <input type="number" class="paymentAmount" placeholder="Enter amount">
                        </div>
                        <div class="form-group required">
                            <label>Date:</label>
                            <input type="date" class="paymentDate">
                        </div>
                    </div>
                    <div class="form-group">
                        <label>Bank:</label>
                        <input type="text" value="SBI" readonly style="background: #f8f9fa;">
                    </div>
                `;
            } else if (mode === 'DD') {
                html = `
                    <div class="form-group required">
                        <label>DD Number:</label>
                        <input type="text" class="paymentRef" placeholder="Enter DD number">
                    </div>
                    <div class="form-row">
                        <div class="form-group required">
                            <label>Amount (₹):</label>
                            <input type="number" class="paymentAmount" placeholder="Enter amount">
                        </div>
                        <div class="form-group required">
                            <label>Date:</label>
                            <input type="date" class="paymentDate">
                        </div>
                    </div>
                    <div class="form-group required">
                        <label>Bank Name:</label>
                        <input type="text" class="paymentBank" placeholder="Enter bank name">
                    </div>
                `;
            }
            
            fieldsElement.innerHTML = html;
        }
        
        function updateRequiredDocs(formType) {
            let certType = '';
            let docsDiv = '';
            
            if (formType === 'regular') {
                certType = document.getElementById('regCertificateType')?.value || '';
                docsDiv = 'regDocsCheckbox';
            } else if (formType === 'distance') {
                certType = document.getElementById('distCertificateType')?.value || '';
                docsDiv = 'distDocsCheckbox';
            }
            
            const docsElement = document.getElementById(docsDiv);
            if (!docsElement) return;
            
            const docs = requiredDocs[certType] || [];
            let html = '';
            docs.forEach(doc => {
                html += `<div class="checkbox-item"><input type="checkbox" value="${doc}"><label style="margin-bottom: 0;">${doc}</label></div>`;
            });
            docsElement.innerHTML = html;
        }
        
        function submitRegularForm() {
            const name = document.getElementById('regName')?.value.trim() || '';
            const register = document.getElementById('regRegister')?.value.trim() || '';
            const hallTicket = document.getElementById('regHallTicket')?.value.trim() || '';
            const address = document.getElementById('regAddress')?.value.trim() || '';
            const degree = document.getElementById('regDegree')?.value || '';
            const course = document.getElementById('regCourse')?.value || '';
            const certificate = document.getElementById('regCertificateType')?.value || '';
            const paymentMode = document.getElementById('regPaymentMode')?.value || '';
            
            if (!name || !register || !hallTicket || !degree || !course || !certificate || !paymentMode) {
                showAlert('regularAlert', '⚠ Please fill all required fields', 'warning');
                return;
            }
            
            // Validate UG specific fields
            let tableNumber = '';
            let collegeCode = '';
            
            if (degree === 'ug') {
                tableNumber = document.getElementById('regTableNumber')?.value || '';
                collegeCode = document.getElementById('regCollege')?.value || '';
                
                if (!tableNumber || !collegeCode) {
                    showAlert('regularAlert', '⚠ Please select table number and college for UG programs', 'warning');
                    return;
                }
            } else if (degree === 'pg' || degree.value === 'mphil') {
                collegeCode = document.getElementById('regPGCollege')?.value.trim() || '';
                tableNumber = document.getElementById('regPGTableNumber')?.value || '';
                
                if (!collegeCode) {
                    showAlert('regularAlert', '⚠ Please enter college code for PG/MPhil programs', 'warning');
                    return;
                }
                if (!tableNumber) {
                    showAlert('regularAlert', '⚠ Table number is not auto-assigned for this course. Please check course selection.', 'warning');
                    return;
                }
            }
            
            const modeElement = document.querySelector('input[name="regMode"]:checked');
            const mode = modeElement ? modeElement.value : 'Original';
            
            const paymentFields = document.querySelectorAll('#regPaymentFields input');
            const paymentRef = paymentFields[0]?.value || '';
            const paymentAmount = paymentFields[1]?.value || '';
            const paymentDate = paymentFields[2]?.value || '';
            
            const data = {
                formType: 'regular',
                data: {
                    name, register, hallticket: hallTicket, address,
                    degree, course, collegeCode, tableNumber,
                    certificateType: certificate, mode, paymentMode, 
                    paymentRef, amount: paymentAmount, paymentDate
                }
            };
            
            fetch('/save-application', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify(data)
            })
            .then(r => r.json())
            .then(data => {
                if (data.success) {
                    showAlert('regularAlert', '✅ Application saved successfully! Form has been reset.', 'success');
                    resetForm('regular');
                    setTimeout(() => {
                        document.getElementById('regularAlert').classList.remove('show');
                    }, 3000);
                } else {
                    showAlert('regularAlert', '❌ Error: ' + (data.error || 'Failed to save application'), 'error');
                }
            })
            .catch(err => {
                showAlert('regularAlert', '❌ Network error: ' + err.message, 'error');
            });
        }
        
        function submitDistanceForm() {
            const name = document.getElementById('distName')?.value.trim() || '';
            const register = document.getElementById('distRegister')?.value.trim() || '';
            const hallTicket = document.getElementById('distHallTicket')?.value.trim() || '';
            const address = document.getElementById('distAddress')?.value.trim() || '';
            const degree = document.getElementById('distDegree')?.value || '';
            const course = document.getElementById('distCourse')?.value || '';
            const certificate = document.getElementById('distCertificateType')?.value || '';
            const paymentMode = document.getElementById('distPaymentMode')?.value || '';
            
            if (!name || !register || !hallTicket || !degree || !course || !certificate || !paymentMode) {
                showAlert('distanceAlert', '⚠ Please fill all required fields', 'warning');
                return;
            }
            
            const modeElement = document.querySelector('input[name="distMode"]:checked');
            const mode = modeElement ? modeElement.value : 'Original';
            
            const paymentFields = document.querySelectorAll('#distPaymentFields input');
            const paymentRef = paymentFields[0]?.value || '';
            const paymentAmount = paymentFields[1]?.value || '';
            const paymentDate = paymentFields[2]?.value || '';
            
            const data = {
                formType: 'distance',
                data: {
                    name, register, hallticket: hallTicket, address,
                    degree, course, tableNumber: 'D-1', 
                    certificateType: certificate, mode, paymentMode, 
                    paymentRef, amount: paymentAmount, paymentDate
                }
            };
            
            fetch('/save-application', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify(data)
            })
            .then(r => r.json())
            .then(data => {
                if (data.success) {
                    showAlert('distanceAlert', '✅ Distance Application saved successfully! Form has been reset.', 'success');
                    resetForm('distance');
                    setTimeout(() => {
                        document.getElementById('distanceAlert').classList.remove('show');
                    }, 3000);
                } else {
                    showAlert('distanceAlert', '❌ Error: ' + (data.error || 'Failed to save application'), 'error');
                }
            })
            .catch(err => {
                showAlert('distanceAlert', '❌ Network error: ' + err.message, 'error');
            });
        }
        
        function submitConfidentialForm() {
            const college = document.getElementById('confCollege')?.value.trim() || '';
            const hallTicket = document.getElementById('confHallTicket')?.value.trim() || '';
            const details = document.getElementById('confDetails')?.value.trim() || '';
            const paymentMode = document.getElementById('confPaymentMode')?.value || '';
            
            if (!college || !hallTicket || !details || !paymentMode) {
                showAlert('confAlert', '⚠ Please fill all required fields', 'warning');
                return;
            }
            
            const data = {
                formType: 'confidential',
                data: {
                    collegeCode: college, 
                    hallticket: hallTicket, 
                    applicationDetails: details,
                    tableNumber: 'C-1', 
                    paymentMode
                }
            };
            
            fetch('/save-application', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify(data)
            })
            .then(r => r.json())
            .then(data => {
                if (data.success) {
                    showAlert('confAlert', '✅ Confidential Application saved successfully! Form has been reset.', 'success');
                    resetForm('confidential');
                    setTimeout(() => {
                        document.getElementById('confAlert').classList.remove('show');
                    }, 3000);
                } else {
                    showAlert('confAlert', '❌ Error: ' + (data.error || 'Failed to save application'), 'error');
                }
            })
            .catch(err => {
                showAlert('confAlert', '❌ Network error: ' + err.message, 'error');
            });
        }
        
        // ... rest of the JavaScript functions remain the same ...
        // [The remaining JavaScript functions for dispatch, admin, etc. remain unchanged]
        
        function showDispatchSection(section) {
            document.getElementById('dispatchProcessing').style.display = 'none';
            document.getElementById('dispatchClosure').style.display = 'none';
            resetDispatchForm();
            
            if (section === 'processing') {
                document.getElementById('dispatchProcessing').style.display = 'block';
            } else if (section === 'closure') {
                document.getElementById('dispatchClosure').style.display = 'block';
            }
        }
        
        function verifyDispatchHallTicket() {
            const hallTicket = document.getElementById('dispatchHallTicket')?.value.trim() || '';
            if (!hallTicket) {
                alert('⚠️ Please enter hall ticket number');
                return;
            }
            
            // Get pending applications for this hall ticket
            fetch('/get-pending-applications', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify({hallticket: hallTicket})
            })
            .then(r => r.json())
            .then(data => {
                if (data.success && data.pending_apps && data.pending_apps.length > 0) {
                    // Show list of pending applications
                    displayPendingApplications(data.pending_apps, hallTicket);
                } else {
                    alert('❌ No pending applications found for this hall ticket');
                    document.getElementById('dispatchHallTicket').value = '';
                }
            })
            .catch(err => {
                alert('❌ Error checking applications: ' + err.message);
            });
        }
        
        function displayPendingApplications(applications, hallTicket) {
            const container = document.getElementById('pendingApplications');
            container.innerHTML = '';
            
            applications.forEach(app => {
                const appDate = new Date(app.submittedAt).toLocaleDateString();
                const appItem = document.createElement('div');
                appItem.className = 'application-item';
                appItem.innerHTML = `
                    <div class="application-details">
                        <div class="application-info">
                            <strong>${app.name}</strong><br>
                            <small>Certificate: ${app.certificateType}</small><br>
                            <small>Submitted: ${appDate}</small><br>
                            <small>Form Type: ${app.formType}</small>
                        </div>
                        <div class="application-actions">
                            <button class="btn btn-primary" onclick="selectApplication(${app.id}, '${hallTicket}')">
                                Select
                            </button>
                        </div>
                    </div>
                `;
                container.appendChild(appItem);
            });
            
            document.getElementById('pendingApplicationsList').style.display = 'block';
        }
        
        function selectApplication(appId, hallTicket) {
            selectedAppId = appId;
            
            // Remove selection from all items
            document.querySelectorAll('.application-item').forEach(item => {
                item.classList.remove('selected');
            });
            
            // Add selection to clicked item
            event.target.closest('.application-item').classList.add('selected');
            
            // Show dispatch details and mode selection
            document.getElementById('dispatchDetails').innerHTML = `
                <strong>✅ Application Selected</strong><br>
                <strong>📋 Application ID:</strong> ${appId}<br>
                <strong>🎫 Hall Ticket:</strong> ${hallTicket}
            `;
            document.getElementById('dispatchDetails').style.display = 'block';
            document.getElementById('dispatchModeGroup').style.display = 'block';
            document.getElementById('dispatchSaveGroup').style.display = 'flex';
        }
        
        function verifyClosureHallTicket() {
            const hallTicket = document.getElementById('closureHallTicket')?.value.trim() || '';
            if (!hallTicket) {
                alert('⚠️ Please enter hall ticket number');
                return;
            }
            
            // Get processed dispatches for this hall ticket
            fetch('/get-processed-dispatches', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify({hallticket: hallTicket})
            })
            .then(r => r.json())
            .then(data => {
                if (data.success && data.processed_dispatches && data.processed_dispatches.length > 0) {
                    // Show list of processed dispatches
                    displayProcessedDispatches(data.processed_dispatches, hallTicket);
                } else {
                    alert('❌ No processed dispatches found for this hall ticket');
                    document.getElementById('closureHallTicket').value = '';
                }
            })
            .catch(err => {
                alert('❌ Error checking dispatches: ' + err.message);
            });
        }
        
        function displayProcessedDispatches(dispatches, hallTicket) {
            const container = document.getElementById('processedDispatches');
            container.innerHTML = '';
            
            dispatches.forEach(dispatch => {
                const dispatchDate = new Date(dispatch.timestamp).toLocaleDateString();
                const dispatchItem = document.createElement('div');
                dispatchItem.className = 'application-item';
                dispatchItem.innerHTML = `
                    <div class="application-details">
                        <div class="application-info">
                            <strong>${dispatch.name}</strong><br>
                            <small>Certificate: ${dispatch.certificateType}</small><br>
                            <small>Dispatch Mode: ${dispatch.dispatchMode}</small><br>
                            <small>Processed: ${dispatchDate}</small>
                        </div>
                        <div class="application-actions">
                            <button class="btn btn-primary" onclick="selectDispatchForClosure(${dispatch.app_id}, '${hallTicket}')">
                                Select
                            </button>
                        </div>
                    </div>
                `;
                container.appendChild(dispatchItem);
            });
            
            document.getElementById('processedDispatchesList').style.display = 'block';
        }
        
        function selectDispatchForClosure(appId, hallTicket) {
            selectedDispatchAppId = appId;
            
            // Remove selection from all items
            document.querySelectorAll('.application-item').forEach(item => {
                item.classList.remove('selected');
            });
            
            // Add selection to clicked item
            event.target.closest('.application-item').classList.add('selected');
            
            // Show closure details and fields
            document.getElementById('closureDetails').innerHTML = `
                <strong>✅ Dispatch Selected for Closure</strong><br>
                <strong>📋 Application ID:</strong> ${appId}<br>
                <strong>🎫 Hall Ticket:</strong> ${hallTicket}
            `;
            document.getElementById('closureDetails').style.display = 'block';
            document.getElementById('closureFields').style.display = 'block';
        }
        
        function saveDispatch() {
            const hallTicket = document.getElementById('dispatchHallTicket')?.value.trim() || '';
            const mode = document.getElementById('dispatchMode')?.value || '';
            
            if (!mode) {
                alert('⚠️ Please select dispatch mode');
                return;
            }
            
            if (!selectedAppId) {
                alert('⚠️ Please select an application to dispatch');
                return;
            }
            
            console.log('Processing dispatch for hall ticket:', hallTicket, 'with mode:', mode, 'app_id:', selectedAppId);
            
            fetch('/save-dispatch', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify({
                    hallticket: hallTicket, 
                    dispatchMode: mode,
                    app_id: selectedAppId
                })
            })
            .then(r => {
                if (!r.ok) {
                    return r.json().then(errData => {
                        throw new Error(errData.error || `HTTP error! status: ${r.status}`);
                    });
                }
                return r.json();
            })
            .then(data => {
                if (data.success) {
                    alert('✅ Dispatch processed successfully! Application is ready for closure.');
                    resetDispatchForm();
                    document.getElementById('dispatchProcessing').style.display = 'none';
                } else {
                    throw new Error(data.error || 'Unknown error');
                }
            })
            .catch(err => {
                console.error('Error:', err);
                alert('❌ Error processing dispatch: ' + err.message);
            });
        }
        
        function closeDispatch() {
            const hallTicket = document.getElementById('closureHallTicket')?.value.trim() || '';
            const trackingId = document.getElementById('closureTrackingId')?.value.trim() || '';
            
            if (!trackingId) {
                alert('⚠️ Please enter tracking ID or reason');
                return;
            }
            
            if (!selectedDispatchAppId) {
                alert('⚠️ Please select a dispatch to close');
                return;
            }
            
            console.log('Closing dispatch for hall ticket:', hallTicket, 'with tracking:', trackingId, 'app_id:', selectedDispatchAppId);
            
            fetch('/close-dispatch', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify({
                    hallticket: hallTicket, 
                    tracking_id: trackingId,
                    app_id: selectedDispatchAppId
                })
            })
            .then(r => {
                if (!r.ok) {
                    return r.json().then(errData => {
                        throw new Error(errData.error || `HTTP error! status: ${r.status}`);
                    });
                }
                return r.json();
            })
            .then(data => {
                if (data.success) {
                    alert('✅ Dispatch closed successfully! Application marked as CLOSED.');
                    resetDispatchForm();
                    document.getElementById('dispatchClosure').style.display = 'none';
                    if (isAdminLoggedIn) {
                        loadAdminData();
                    }
                } else {
                    throw new Error(data.error || 'Unknown error');
                }
            })
            .catch(err => {
                console.error('Error:', err);
                alert('❌ Error closing dispatch: ' + err.message);
            });
        }
        
        function adminLogin() {
            const username = document.getElementById('adminUsername')?.value || '';
            const password = document.getElementById('adminPassword')?.value || '';
            
            if (username === 'admin' && password === 'pass') {
                isAdminLoggedIn = true;
                document.getElementById('adminLogin').style.display = 'none';
                document.getElementById('adminDashboard').style.display = 'block';
                loadAdminData();
                autoRefreshInterval = setInterval(loadAdminData, 10000);
            } else {
                showAlert('loginAlert', '❌ Invalid credentials', 'error');
            }
        }
        
        function loadAdminData() {
            fetch('/get-all-data')
                .then(r => r.json())
                .then(data => {
                    updateStatistics(data.applications || []);
                })
                .catch(err => {
                    console.error('Error loading admin data:', err);
                });
        }
        
        function updateStatistics(apps) {
            const today = new Date().toDateString();
            let todayCount = 0, closedCount = 0, pendingCount = 0, withinCount = 0, outofCount = 0, overdueCount = 0;
            
            apps.forEach(app => {
                try {
                    const appDate = new Date(app.submittedAt).toDateString();
                    if (appDate === today) todayCount++;
                    if (app.status === 'closed') closedCount++;
                    else if (app.status === 'pending') pendingCount++;
                    
                    const status = app.calculatedStatus;
                    if (status === 'within_time') withinCount++;
                    else if (status === 'out_of_time') outofCount++;
                    else if (status === 'overdue') overdueCount++;
                } catch (e) {
                    console.error('Error processing application:', e);
                }
            });
            
            document.getElementById('statToday').textContent = todayCount;
            document.getElementById('statClosed').textContent = closedCount;
            document.getElementById('statPending').textContent = pendingCount;
            document.getElementById('statWithin').textContent = withinCount;
            document.getElementById('statOutOf').textContent = outofCount;
            document.getElementById('statOverdue').textContent = overdueCount;
        }
        
        function exportDateRangeReport() {
            const fromDate = document.getElementById('reportFromDate')?.value || '';
            const toDate = document.getElementById('reportToDate')?.value || '';
            
            if (!fromDate || !toDate) {
                alert('⚠️ Please select both dates');
                return;
            }
            
            fetch('/export-excel', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify({
                    fromDate: fromDate + 'T00:00:00', 
                    toDate: toDate + 'T23:59:59', 
                    tableNumber: null
                })
            })
            .then(r => r.blob())
            .then(blob => {
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `report_${fromDate}_${toDate}.xlsx`;
                a.click();
            })
            .catch(err => {
                alert('❌ Error exporting report: ' + err.message);
            });
        }
        
        function exportTableWiseReport() {
            const table = document.getElementById('reportTableNumber')?.value || '';
            const fromDate = document.getElementById('reportTableFromDate')?.value || '';
            const toDate = document.getElementById('reportTableToDate')?.value || '';
            
            if (!table || !fromDate || !toDate) {
                alert('⚠️ Please select table and both dates');
                return;
            }
            
            fetch('/export-excel', {
                method: 'POST', 
                headers: {'Content-Type': 'application/json'}, 
                body: JSON.stringify({
                    fromDate: fromDate + 'T00:00:00', 
                    toDate: toDate + 'T23:59:59', 
                    tableNumber: table
                })
            })
            .then(r => r.blob())
            .then(blob => {
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `table_${table}_${fromDate}_${toDate}.xlsx`;
                a.click();
            })
            .catch(err => {
                alert('❌ Error exporting report: ' + err.message);
            });
        }
        
        function refreshAdminData() {
            loadAdminData();
            alert('✅ Data refreshed');
        }
        
        function showAlert(elementId, message, type) {
            const alert = document.getElementById(elementId);
            if (alert) {
                alert.textContent = message;
                alert.className = `alert show alert-${type}`;
                setTimeout(() => {
                    alert.classList.remove('show');
                }, 4000);
            }
        }
    </script>
</body>
</html>
"""

if __name__ == '__main__':
    app.run(debug=True, port=5000)